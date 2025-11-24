from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Tuple, Union
import unicodedata
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import math


EXPECTED_FIELDS = {
    "company_name",
    "name",
}

SUPPLIER_EXPECTED_FIELDS = {"name"}

COLUMN_ALIASES: Dict[str, str] = {
    "company_name": "company_name",
    "entreprise": "company_name",
    "nom_entreprise": "company_name",
    "societe": "company_name",
    "raison_sociale": "company_name",
    "name": "name",
    "client": "name",
    "nom_client": "name",
    "contact": "name",
    "email": "email",
    "mail": "email",
    "courriel": "email",
    "phone": "phone",
    "telephone": "phone",
    "tel": "phone",
    "technician_name": "technician_name",
    "technicien": "technician_name",
    "technicien_referent": "technician_name",
    "referent": "technician_name",
    "billing_address": "billing_address",
    "adresse": "billing_address",
    "adresse_facturation": "billing_address",
    "depannage": "depannage",
    "type_depannage": "depannage",
    "astreinte": "astreinte",
    "tags": "tags",
    "tag": "tags",
    "status": "status",
    "statut": "status",
}

SUPPLIER_COLUMN_ALIASES: Dict[str, str] = {
    "name": "name",
    "nom": "name",
    "our_code": "our_code",
    "code": "our_code",
    "code_interne": "our_code",
    "code_fournisseur": "our_code",
    "supplier_type": "supplier_type",
    "type": "supplier_type",
    "type_fournisseur": "supplier_type",
    "categories": "categories",
    "categorie": "categories",
    "category": "categories",
}

CONTACT_FIELD_ALIASES = {
    "name": "name",
    "nom": "name",
    "prenom": "name",
    "nom_complet": "name",
    "email": "email",
    "mail": "email",
    "courriel": "email",
    "phone": "phone",
    "telephone": "phone",
    "tel": "phone",
    "mobile": "phone",
    "description": "description",
    "notes": "description",
    "commentaire": "description",
}

SUPPLIER_TYPE_ALIASES = {
    "fournisseur": "fournisseur",
    "fournisseurs": "fournisseur",
    "sous_traitant": "sous_traitant",
    "sous traitant": "sous_traitant",
    "sous-traitant": "sous_traitant",
}

HeaderType = Union[str, Tuple[str, int, str]]

STATUS_ALIASES = {
    "actif": "actif",
    "active": "actif",
    "oui": "actif",
    "true": "actif",
    "1": "actif",
    "inactif": "inactif",
    "inactive": "inactif",
    "non": "inactif",
    "false": "inactif",
    "0": "inactif",
}

DEPANNAGE_CHOICES = {
    "refacturable",
    "non_refacturable",
}

ASTREINTE_CHOICES = {
    "incluse_non_refacturable",
    "incluse_refacturable",
    "pas_d_astreinte",
}

FREQUENCY_UNIT_ALIASES = {
    "mois": "months",
    "month": "months",
    "months": "months",
    "annee": "years",
    "annees": "years",
    "ans": "years",
    "year": "years",
    "years": "years",
}

BOOLEAN_TRUE_VALUES = {"oui", "o", "yes", "y", "true", "vrai", "1"}
BOOLEAN_FALSE_VALUES = {"non", "no", "n", "false", "faux", "0"}

FILTER_EXPECTED_FIELDS = {"site", "equipment", "format_type"}

FILTER_COLUMN_ALIASES: Dict[str, str] = {
    "site": "site",
    "lieu": "site",
    "equipement": "equipment",
    "equipment": "equipment",
    "machine": "equipment",
    "format": "format_type",
    "format_type": "format_type",
    "type": "format_type",
    "efficacite": "efficiency",
    "efficiency": "efficiency",
    "classe": "efficiency",
    "dimensions": "dimensions",
    "dimension": "dimensions",
    "taille": "dimensions",
    "quantite": "quantity",
    "quantity": "quantity",
    "qte": "quantity",
    "commande": "order_week",
    "semaine_commande": "order_week",
    "order_week": "order_week",
    "semaine": "order_week",
    "inclus": "included_in_contract",
    "inclus_contrat": "included_in_contract",
    "included": "included_in_contract",
    "included_in_contract": "included_in_contract",
    "etat_commande": "ordered",
    "statut_commande": "ordered",
    "commande_effectuee": "ordered",
    "ordered": "ordered",
    "poches": "pocket_count",
    "nb_poches": "pocket_count",
    "pocket_count": "pocket_count",
}

FILTER_FORMAT_LOOKUP = {
    "cousus_sur_fil": "cousus_sur_fil",
    "format_cousus_sur_fil": "cousus_sur_fil",
    "cadre": "cadre",
    "format_cadre": "cadre",
    "multi_diedre": "multi_diedre",
    "format_multi_diedre": "multi_diedre",
    "poche": "poche",
    "format_poche": "poche",
}

BELT_EXPECTED_FIELDS = {"site", "equipment", "reference"}

BELT_COLUMN_ALIASES: Dict[str, str] = {
    "site": "site",
    "lieu": "site",
    "equipement": "equipment",
    "equipment": "equipment",
    "machine": "equipment",
    "reference": "reference",
    "ref": "reference",
    "code": "reference",
    "quantite": "quantity",
    "quantity": "quantity",
    "qte": "quantity",
    "commande": "order_week",
    "semaine_commande": "order_week",
    "order_week": "order_week",
    "semaine": "order_week",
    "inclus": "included_in_contract",
    "inclus_contrat": "included_in_contract",
    "included": "included_in_contract",
    "included_in_contract": "included_in_contract",
    "etat_commande": "ordered",
    "statut_commande": "ordered",
    "commande_effectuee": "ordered",
    "ordered": "ordered",
}

PRESTATION_COLUMN_ALIASES: Dict[str, str] = {
    "company_name": "company_name",
    "entreprise": "company_name",
    "societe": "company_name",
    "raison_sociale": "company_name",
    "client": "client_name",
    "client_name": "client_name",
    "contact": "client_name",
    "nom_client": "client_name",
    "prestation": "prestation",
    "prestation_key": "prestation",
    "libelle": "prestation_label",
    "prestation_label": "prestation_label",
    "budget": "budget",
    "montant": "budget",
    "frequency": "frequency",
    "frequence": "frequency",
    "intervalle": "frequency_interval",
    "frequency_interval": "frequency_interval",
    "frequency_unit": "frequency_unit",
    "unite": "frequency_unit",
    "client_id": "client_id",
    "status": "status",
    "statut": "status",
    "order_week": "order_week",
    "commande": "order_week",
    "semaine_commande": "order_week",
    "realization_week": "realization_week",
    "realisation_week": "realization_week",
    "semaine_realisation": "realization_week",
}


WORKLOAD_STATE_ALIASES = {
    "warn": "warn",
    "warning": "warn",
    "orange": "warn",
    "attention": "warn",
    "a": "warn",
    "bad": "bad",
    "rouge": "bad",
    "red": "bad",
    "critique": "bad",
    "r": "bad",
    "ok": "ok",
    "vert": "ok",
    "green": "ok",
    "g": "ok",
}


def _normalize_header(header: Optional[str]) -> str:
    if header is None:
        return ""
    value = str(header).strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    for char in ("-", ".", "'", "\u00a0"):
        value = value.replace(char, " ")
    return "_".join(part for part in value.split() if part)


def _coerce_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).rstrip("0").rstrip(".")
        return str(value)
    return str(value).strip() or None


CONTACT_HEADER_RE = re.compile(r"contact_?(\d+)_([a-z0-9_]+)")


def _resolve_header(header: str) -> HeaderType:
    if header in COLUMN_ALIASES:
        return COLUMN_ALIASES[header]
    match = CONTACT_HEADER_RE.match(header)
    if match:
        index = int(match.group(1))
        field_key = match.group(2)
        field = CONTACT_FIELD_ALIASES.get(field_key)
        if field:
            return ("contact", index, field)
    return ""


def _resolve_supplier_header(header: str) -> HeaderType:
    if header in SUPPLIER_COLUMN_ALIASES:
        return SUPPLIER_COLUMN_ALIASES[header]
    match = CONTACT_HEADER_RE.match(header)
    if match:
        index = int(match.group(1))
        field_key = match.group(2)
        field = CONTACT_FIELD_ALIASES.get(field_key)
        if field:
            return ("contact", index, field)
    return ""


def _resolve_filter_header(header: str) -> str:
    return FILTER_COLUMN_ALIASES.get(header, "")


def _resolve_belt_header(header: str) -> str:
    return BELT_COLUMN_ALIASES.get(header, "")


def _resolve_prestation_header(header: str) -> str:
    return PRESTATION_COLUMN_ALIASES.get(header, "")


def _parse_positive_int(value: str, row_index: int, field_name: str) -> int:
    try:
        if "." in value:
            parsed = float(value)
            if not parsed.is_integer():
                raise ValueError
            quantity = int(parsed)
        else:
            quantity = int(value)
    except ValueError:
        raise ValueError(
            f"Ligne {row_index}: valeur de {field_name} invalide '{value}'."
        ) from None

    if quantity < 1:
        raise ValueError(
            f"Ligne {row_index}: valeur de {field_name} invalide '{value}'."
        )
    return quantity


def _parse_boolean_flag(value: str, row_index: int, field_name: str) -> bool:
    normalized = _normalize_header(value)
    if normalized in BOOLEAN_TRUE_VALUES:
        return True
    if normalized in BOOLEAN_FALSE_VALUES:
        return False
    raise ValueError(
        f"Ligne {row_index}: valeur de {field_name} invalide '{value}'."
    )


def parse_clients_excel(content: bytes) -> List[Dict[str, str]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # pragma: no cover - delegated to openpyxl
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        raise ValueError("Le fichier ne contient aucune donnée.")

    headers: List[HeaderType] = [
        _resolve_header(_normalize_header(cell.value)) for cell in header_row
    ]

    if not any(headers):
        raise ValueError("Le fichier ne contient pas d'en-têtes valides.")

    missing = EXPECTED_FIELDS - {
        header for header in headers if isinstance(header, str) and header
    }
    if missing:
        raise ValueError(
            "Colonnes obligatoires manquantes: " + ", ".join(sorted(missing))
        )

    rows: List[Dict[str, str]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record: Dict[str, str] = {"__row__": row_index}
        empty = True
        for idx, raw_value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            value = _coerce_value(raw_value)
            if value is None:
                continue
            empty = False
            if isinstance(header, tuple):
                _, contact_index, contact_field = header
                contact_bucket = record.setdefault("__contacts__", {})
                contact_data = contact_bucket.setdefault(contact_index, {})
                contact_data[contact_field] = value
                continue

            key = header
            if key == "status":
                normalized = _normalize_header(value)
                record[key] = STATUS_ALIASES.get(normalized, normalized or None)
            elif key == "depannage":
                normalized = _normalize_header(value)
                if normalized and normalized not in DEPANNAGE_CHOICES:
                    raise ValueError(
                        f"Ligne {row_index}: valeur de dépannage invalide '{value}'."
                    )
                if normalized:
                    record[key] = normalized
            elif key == "astreinte":
                normalized = _normalize_header(value)
                if normalized and normalized not in ASTREINTE_CHOICES:
                    raise ValueError(
                        f"Ligne {row_index}: valeur d'astreinte invalide '{value}'."
                    )
                if normalized:
                    record[key] = normalized
            else:
                record[key] = value

        if empty:
            continue

        missing_fields = [field for field in EXPECTED_FIELDS if not record.get(field)]
        if missing_fields:
            raise ValueError(
                f"Ligne {row_index}: valeurs manquantes pour {', '.join(missing_fields)}"
            )

        if record.get("status") and record["status"] not in STATUS_ALIASES.values():
            raise ValueError(
                f"Ligne {row_index}: statut inconnu '{record['status']}'. Valeurs acceptées: actif, inactif."
            )

        contacts_map = record.pop("__contacts__", {})
        contacts: List[Dict[str, str]] = []
        for order, data in sorted(contacts_map.items()):
            if not data.get("name"):
                if any(data.get(field) for field in ("email", "phone")):
                    raise ValueError(
                        f"Ligne {row_index}: le contact {order} doit avoir un nom."
                    )
                continue
            contacts.append(data)

        if contacts:
            record["contacts"] = contacts

        rows.append(record)

    return rows


def parse_suppliers_excel(content: bytes) -> List[Dict[str, str]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # pragma: no cover - delegated to openpyxl
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        raise ValueError("Le fichier ne contient aucune donnée.")

    headers: List[HeaderType] = [
        _resolve_supplier_header(_normalize_header(cell.value)) for cell in header_row
    ]

    if not any(headers):
        raise ValueError("Le fichier ne contient pas d'en-têtes valides.")

    missing = SUPPLIER_EXPECTED_FIELDS - {
        header for header in headers if isinstance(header, str) and header
    }
    if missing:
        raise ValueError(
            "Colonnes obligatoires manquantes: " + ", ".join(sorted(missing))
        )

    rows: List[Dict[str, str]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record: Dict[str, str] = {"__row__": row_index}
        empty = True
        for idx, raw_value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            value = _coerce_value(raw_value)
            if value is None:
                continue
            empty = False
            if isinstance(header, tuple):
                _, contact_index, contact_field = header
                contact_bucket = record.setdefault("__contacts__", {})
                contact_data = contact_bucket.setdefault(contact_index, {})
                contact_data[contact_field] = value
                continue

            key = header
            if key == "supplier_type":
                normalized = _normalize_header(value)
                record[key] = SUPPLIER_TYPE_ALIASES.get(normalized)
                if value and not record[key]:
                    raise ValueError(
                        f"Ligne {row_index}: type de fournisseur inconnu '{value}'."
                    )
            else:
                record[key] = value

        if empty:
            continue

        missing_fields = [
            field for field in SUPPLIER_EXPECTED_FIELDS if not record.get(field)
        ]
        if missing_fields:
            raise ValueError(
                f"Ligne {row_index}: valeurs manquantes pour {', '.join(missing_fields)}"
            )

        contacts_map = record.pop("__contacts__", {})
        contacts: List[Dict[str, str]] = []
        for order, data in sorted(contacts_map.items()):
            if not data.get("name"):
                if any(data.get(field) for field in ("email", "phone")):
                    raise ValueError(
                        f"Ligne {row_index}: le contact {order} doit avoir un nom."
                    )
                continue
            contacts.append(data)

        if contacts:
            record["contacts"] = contacts

        rows.append(record)

    return rows


def parse_prestations_excel(content: bytes) -> List[Dict[str, Union[str, int]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        raise ValueError("Le fichier ne contient aucune donnée.")

    headers = [
        _resolve_prestation_header(_normalize_header(cell.value)) for cell in header_row
    ]

    if not any(headers):
        raise ValueError("Le fichier ne contient pas d'en-têtes valides.")

    rows: List[Dict[str, Union[str, int]]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record: Dict[str, Union[str, int]] = {"__row__": row_index}
        empty = True
        for idx, raw_value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            value = _coerce_value(raw_value)
            if value is None:
                continue
            empty = False
            if header in {"frequency", "status"}:
                record[header] = _normalize_header(value)
            elif header == "frequency_unit":
                normalized_unit = _normalize_header(value)
                record[header] = FREQUENCY_UNIT_ALIASES.get(
                    normalized_unit, normalized_unit
                )
            elif header in {"company_name", "client_name", "prestation_label", "prestation"}:
                record[header] = value.strip()
            elif header == "client_id":
                record[header] = _parse_positive_int(value, row_index, "identifiant client")
            elif header == "frequency_interval":
                record[header] = _parse_positive_int(
                    value, row_index, "intervalle de fréquence"
                )
            else:
                record[header] = value

        if empty:
            continue

        if not record.get("prestation") and not record.get("prestation_label"):
            raise ValueError(
                f"Ligne {row_index}: veuillez renseigner la colonne prestation ou libellé."
            )

        has_client_reference = any(
            record.get(field)
            for field in ("client_id", "company_name", "client_name")
        )
        if not has_client_reference:
            raise ValueError(
                "Ligne {row_index}: renseignez le nom d'entreprise, le contact client ou l'identifiant client."
                .format(row_index=row_index)
            )

        rows.append(record)

    return rows


def parse_filter_lines_excel(content: bytes) -> List[Dict[str, Union[str, int]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        raise ValueError("Le fichier ne contient aucune donnée.")

    headers = [
        _resolve_filter_header(_normalize_header(cell.value)) for cell in header_row
    ]

    if not any(headers):
        raise ValueError("Le fichier ne contient pas d'en-têtes valides.")

    missing = FILTER_EXPECTED_FIELDS - {header for header in headers if header}
    if missing:
        raise ValueError(
            "Colonnes obligatoires manquantes: " + ", ".join(sorted(missing))
        )

    rows: List[Dict[str, Union[str, int]]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record: Dict[str, Union[str, int]] = {"__row__": row_index}
        empty = True
        for idx, raw_value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            value = _coerce_value(raw_value)
            if value is None:
                continue
            empty = False
            if header == "format_type":
                normalized = _normalize_header(value)
                resolved = FILTER_FORMAT_LOOKUP.get(normalized)
                if not resolved:
                    raise ValueError(
                        f"Ligne {row_index}: format de filtre inconnu '{value}'."
                    )
                record[header] = resolved
            elif header == "quantity":
                record[header] = _parse_positive_int(value, row_index, "quantité")
            elif header == "pocket_count":
                record[header] = _parse_positive_int(value, row_index, "nombre de poches")
            elif header == "order_week":
                record[header] = value.strip().upper()
            elif header == "included_in_contract":
                record[header] = _parse_boolean_flag(
                    value, row_index, "inclus au contrat"
                )
            elif header == "ordered":
                record[header] = _parse_boolean_flag(
                    value, row_index, "commandé"
                )
            else:
                record[header] = value

        if empty:
            continue

        missing_fields = [
            field for field in FILTER_EXPECTED_FIELDS if not record.get(field)
        ]
        if missing_fields:
            raise ValueError(
                "Ligne {row_index}: champ(s) obligatoire(s) manquant(s): "
                + ", ".join(sorted(missing_fields))
            )

        if "quantity" not in record:
            record["quantity"] = 1

        if "included_in_contract" not in record:
            record["included_in_contract"] = False

        if "ordered" not in record:
            record["ordered"] = False

        if record.get("format_type") == "poche" and "pocket_count" not in record:
            raise ValueError(
                f"Ligne {row_index}: indiquez le nombre de poches pour un filtre au format poche."
            )

        if record.get("format_type") != "poche":
            record.pop("pocket_count", None)

        rows.append(record)

    return rows


def parse_belt_lines_excel(content: bytes) -> List[Dict[str, Union[str, int]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        raise ValueError("Le fichier ne contient aucune donnée.")

    headers = [
        _resolve_belt_header(_normalize_header(cell.value)) for cell in header_row
    ]

    if not any(headers):
        raise ValueError("Le fichier ne contient pas d'en-têtes valides.")

    missing = BELT_EXPECTED_FIELDS - {header for header in headers if header}
    if missing:
        raise ValueError(
            "Colonnes obligatoires manquantes: " + ", ".join(sorted(missing))
        )

    rows: List[Dict[str, Union[str, int]]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record: Dict[str, Union[str, int]] = {"__row__": row_index}
        empty = True
        for idx, raw_value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            value = _coerce_value(raw_value)
            if value is None:
                continue
            empty = False
            if header == "quantity":
                record[header] = _parse_positive_int(value, row_index, "quantité")
            elif header == "order_week":
                record[header] = value.strip().upper()
            elif header == "included_in_contract":
                record[header] = _parse_boolean_flag(
                    value, row_index, "inclus au contrat"
                )
            elif header == "ordered":
                record[header] = _parse_boolean_flag(
                    value, row_index, "commandé"
                )
            else:
                record[header] = value

        if empty:
            continue

        missing_fields = [
            field for field in BELT_EXPECTED_FIELDS if not record.get(field)
        ]
        if missing_fields:
            raise ValueError(
                "Ligne {row_index}: champ(s) obligatoire(s) manquant(s): "
                + ", ".join(sorted(missing_fields))
            )

        if "quantity" not in record:
            record["quantity"] = 1

        if "included_in_contract" not in record:
            record["included_in_contract"] = False

        if "ordered" not in record:
            record["ordered"] = False

        rows.append(record)

    return rows


def _format_hours(value: float) -> str:
    if math.isnan(value) or math.isinf(value):
        raise ValueError("Durée invalide")
    rounded = round(value, 4)
    if rounded.is_integer():
        return str(int(rounded))
    text = f"{rounded:.2f}".rstrip("0").rstrip(".")
    return text or str(rounded)


def _parse_hours_value(raw: str, row_index: int, column_index: int) -> str:
    cleaned = raw.strip().lower()
    for suffix in ("heures", "hours", "heure", "hour", "h"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    cleaned = cleaned.strip().replace(",", ".")
    if not cleaned:
        raise ValueError(_format_workload_error(raw, row_index, column_index))
    try:
        value = float(cleaned)
    except ValueError:
        raise ValueError(_format_workload_error(raw, row_index, column_index)) from None
    if value <= 0:
        raise ValueError(_format_workload_error(raw, row_index, column_index))
    return _format_hours(value)


def _format_workload_error(value: str, row_index: int, column_index: int) -> str:
    column_letter = get_column_letter(column_index)
    return (
        f"Ligne {row_index}, colonne {column_letter}: valeur '{value}' invalide. "
        "Utilisez bad, warn ou ok:4."
    )


def _normalize_workload_cell_value(
    raw_value: Optional[Union[str, int, float]],
    row_index: int,
    column_index: int,
) -> Optional[str]:
    if raw_value is None:
        return None
    if isinstance(raw_value, float) and math.isnan(raw_value):
        return None
    text = str(raw_value).strip()
    if not text:
        return None
    lower = text.lower()

    if lower in WORKLOAD_STATE_ALIASES:
        return WORKLOAD_STATE_ALIASES[lower]

    if lower in {"4", "4h"}:
        return "warn"
    if lower in {"8", "8h"}:
        return "bad"

    if lower.startswith("ok"):
        remainder = lower[2:]
        remainder = remainder.lstrip(" :_-")
        if not remainder:
            return "ok"
        return f"ok:{_parse_hours_value(remainder, row_index, column_index)}"

    if ":" in lower:
        prefix, suffix = lower.split(":", 1)
        prefix = prefix.strip()
        suffix = suffix.strip()
        if prefix in WORKLOAD_STATE_ALIASES:
            state = WORKLOAD_STATE_ALIASES[prefix]
            if state == "ok":
                return f"ok:{_parse_hours_value(suffix, row_index, column_index)}"
            return state

    numeric_match = re.match(r"^(\d+(?:[.,]\d+)?)\s*(h|heures|hours|heure|hour)?$", lower)
    if numeric_match:
        number = float(numeric_match.group(1).replace(",", "."))
        if number <= 0:
            raise ValueError(_format_workload_error(text, row_index, column_index))
        if number == 4:
            return "warn"
        if number == 8:
            return "bad"
        return f"ok:{_format_hours(number)}"

    raise ValueError(_format_workload_error(text, row_index, column_index))


def parse_workload_plan_excel(
    content: bytes,
) -> Tuple[List[str], Dict[str, List[Optional[str]]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel: {exc}")

    sheet = workbook.active
    if sheet.max_row < 2:
        raise ValueError("Le fichier ne contient aucune donnée.")

    sites: List[str] = []
    cells_map: Dict[str, List[Optional[str]]] = {}
    seen_sites: set[str] = set()

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=365, values_only=True), start=2
    ):
        if not row:
            continue
        row_values = list(row)
        if not any(
            (isinstance(value, float) and not math.isnan(value))
            or isinstance(value, int)
            or (isinstance(value, str) and value.strip())
            for value in row_values
        ):
            continue

        raw_site = row_values[0] if len(row_values) > 0 else None
        site_name = _coerce_value(raw_site)
        if not site_name:
            raise ValueError(f"Ligne {row_index}: nom de site manquant.")
        normalized_site = site_name.strip()
        if normalized_site in seen_sites:
            raise ValueError(
                f"Ligne {row_index}: le site '{normalized_site}' est présent plusieurs fois."
            )
        seen_sites.add(normalized_site)
        sites.append(normalized_site)

        cells: List[Optional[str]] = [None] * 364
        for day_index in range(364):
            raw_value = row_values[day_index + 1] if day_index + 1 < len(row_values) else None
            try:
                normalized_value = _normalize_workload_cell_value(
                    raw_value, row_index, day_index + 2
                )
            except ValueError as exc:
                raise ValueError(str(exc))
            cells[day_index] = normalized_value

        cells_map[normalized_site] = cells

    if not sites:
        raise ValueError("Aucun site valide trouvé dans le fichier.")

    return sites, cells_map
