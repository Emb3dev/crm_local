from typing import Optional, Dict, List, Iterable, Union, Annotated, Any, Tuple

from decimal import Decimal

import logging
import os
import secrets
import re
from types import SimpleNamespace
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote

from fastapi import (
    FastAPI,
    Depends,
    Request,
    Form,
    HTTPException,
    UploadFile,
    File,
    status,
)
from fastapi.exception_handlers import http_exception_handler
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlmodel import Session
from jose import JWTError, jwt
from passlib.context import CryptContext

from database import init_db, get_session, engine
from defaults import DEFAULT_PRESTATION_GROUPS
from models import (
    BeltLineCreate,
    BeltLineUpdate,
    ClientCreate,
    ClientUpdate,
    ContactCreate,
    FilterLineCreate,
    FilterLineUpdate,
    SubcontractedServiceCreate,
    SubcontractedServiceUpdate,
    WorkloadCellUpdate,
    WorkloadSiteCreate,
    WorkloadSiteUpdate,
    User,
    UserCreate,
    PrestationDefinitionCreate,
    PrestationDefinitionUpdate,
    PrestationDefinition,
)
import crud
from importers import (
    parse_belt_lines_excel,
    parse_clients_excel,
    parse_filter_lines_excel,
    parse_prestations_excel,
    parse_workload_plan_excel,
)
from openpyxl import Workbook
from uuid import uuid4
from pydantic import BaseModel, Field
#uvicorn app:app --reload

app = FastAPI(title="CRM Local")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

logger = logging.getLogger("crm_local.auth")

SECRET_KEY = os.environ.get("CRM_SECRET_KEY", "change-me")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("CRM_TOKEN_EXPIRE_MINUTES", "480"))
DEFAULT_ADMIN_USERNAME = os.environ.get("CRM_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("CRM_ADMIN_PASSWORD", "admin")

templates.env.globals["ADMIN_USERNAME"] = DEFAULT_ADMIN_USERNAME

MAX_BCRYPT_PASSWORD_BYTES = 72
PASSWORD_MIN_LENGTH = 8

SESSION_COOKIE_NAME = os.environ.get("CRM_SESSION_COOKIE_NAME", "session_token")
SESSION_COOKIE_SECURE = os.environ.get("CRM_SESSION_COOKIE_SECURE", "false").lower() in {"1", "true", "yes"}

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token", auto_error=False)


def _generate_temporary_password(length: int = 12) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


class Token(BaseModel):
    access_token: str
    token_type: str


class TokenData(BaseModel):
    username: Optional[str] = None


def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return pwd_context.verify(plain_password, hashed_password)
    except ValueError:
        return False


def get_password_hash(password: str) -> str:
    if len(password.encode("utf-8")) > MAX_BCRYPT_PASSWORD_BYTES:
        raise ValueError(
            "Le mot de passe dépasse la longueur maximale prise en charge par bcrypt (72 octets)."
        )
    return pwd_context.hash(password)


def authenticate_user(session: Session, username: str, password: str) -> Optional[User]:
    user = crud.get_user_by_username(session, username)
    if not user:
        return None
    if not verify_password(password, user.hashed_password):
        return None
    return user


def create_access_token(
    data: Dict[str, Union[str, int]],
    expires_delta: Optional[timedelta] = None,
) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (
        expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_token_from_request(
    request: Request, token: Optional[str] = Depends(oauth2_scheme)
) -> str:
    if token:
        return token
    cookie_token = request.cookies.get("session_token")
    if cookie_token:
        return cookie_token
    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Authentification requise",
        headers={"WWW-Authenticate": "Bearer"},
    )


@app.exception_handler(HTTPException)
async def redirect_unauthenticated_users(
    request: Request, exc: HTTPException
):
    if exc.status_code == status.HTTP_401_UNAUTHORIZED:
        next_path = request.url.path
        if request.url.query:
            next_path = f"{next_path}?{request.url.query}"
        return RedirectResponse(
            url=f"/login?next={quote(next_path)}",
            status_code=status.HTTP_303_SEE_OTHER,
        )
    return await http_exception_handler(request, exc)


def get_current_user(
    request: Request,
    token: str = Depends(get_token_from_request),
    session: Session = Depends(get_session),
) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Jeton d'authentification invalide",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: Optional[str] = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError as exc:
        raise credentials_exception from exc
    user = crud.get_user_by_username(session, token_data.username or "")
    if not user:
        raise credentials_exception
    crud.touch_user_activity(session, user)
    request.state.user = user
    return user


CurrentUser = Annotated[User, Depends(get_current_user)]


def _ensure_default_admin_user() -> None:
    with Session(engine) as session:
        if crud.get_user_by_username(session, DEFAULT_ADMIN_USERNAME):
            return
        if (
            DEFAULT_ADMIN_PASSWORD == "admin"
            and os.environ.get("CRM_ADMIN_PASSWORD") is None
        ):
            logger.warning(
                "CRM_ADMIN_PASSWORD n'est pas défini : utilisation d'identifiants par défaut 'admin/admin'."
            )
        try:
            hashed_password = get_password_hash(DEFAULT_ADMIN_PASSWORD)
        except ValueError as exc:
            raise RuntimeError(
                "Le mot de passe administrateur par défaut dépasse la limite de 72 octets imposée par bcrypt. "
                "Veuillez définir CRM_ADMIN_PASSWORD avec une valeur plus courte."
            ) from exc
        crud.create_user(
            session,
            UserCreate(username=DEFAULT_ADMIN_USERNAME, hashed_password=hashed_password),
        )
        logger.info(
            "Utilisateur administrateur '%s' initialisé.", DEFAULT_ADMIN_USERNAME
        )


def _ensure_admin_access(user: User) -> None:
    if user.username != DEFAULT_ADMIN_USERNAME:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Accès réservé à l'administrateur.",
        )


DEPANNAGE_OPTIONS = {
    "refacturable": "Refacturable",
    "non_refacturable": "Non refacturable",
}

ASTREINTE_OPTIONS = {
    "incluse_non_refacturable": "Incluse mais non refacturable",
    "incluse_refacturable": "Incluse mais refacturable",
    "pas_d_astreinte": "Pas d'astreinte",
}

STATUS_OPTIONS = {
    "actif": "Actif",
    "inactif": "Inactif",
}

SUBCONTRACT_STATUS_DEFAULT = "non_commence"

SUBCONTRACT_STATUS_OPTIONS = {
    "non_commence": "Non commencé",
    "en_cours": "En cours",
    "fait": "Fait",
}

SUBCONTRACT_STATUS_STYLES = {
    "fait": "border-emerald-200 bg-emerald-50 text-emerald-700 dark:border-emerald-400/40 dark:bg-emerald-500/10 dark:text-emerald-100",
    "en_cours": "border-amber-200 bg-amber-50 text-amber-700 dark:border-amber-400/40 dark:bg-amber-500/10 dark:text-amber-100",
    "non_commence": "border-rose-200 bg-rose-50 text-rose-700 dark:border-rose-400/40 dark:bg-rose-500/10 dark:text-rose-100",
}

FILTER_FORMAT_OPTIONS = [
    ("cousus_sur_fil", "Cousus sur fil"),
    ("cadre", "Format cadre"),
    ("multi_diedre", "Format multi dièdre"),
    ("poche", "Format poche"),
]
FILTER_FORMAT_LABELS = {value: label for value, label in FILTER_FORMAT_OPTIONS}

ALLOWED_IMPORT_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")


def _status_to_bool(value: Optional[str]) -> Optional[bool]:
    if value is None:
        return None
    if value not in STATUS_OPTIONS:
        raise HTTPException(400, "Statut inconnu")
    return value == "actif"


def _status_key_from_bool(value: Optional[bool]) -> str:
    if value is None:
        return ""
    return "actif" if value else "inactif"


def _consume_import_report(request: Request):
    report_token = request.query_params.get("report")
    if not report_token:
        return None
    reports = getattr(app.state, "import_reports", None)
    if reports is None:
        reports = {}
        app.state.import_reports = reports
    return reports.pop(report_token, None)


def _store_import_report(
    redirect_url: str,
    *,
    created: int,
    total: int,
    errors: List[str],
    filename: str,
    singular_label: str,
    plural_label: str,
) -> RedirectResponse:
    report_id = uuid4().hex
    reports = getattr(app.state, "import_reports", None)
    if reports is None:
        reports = {}
        app.state.import_reports = reports
    reports[report_id] = {
        "created": created,
        "errors": errors,
        "total": total,
        "filename": filename,
        "entity_label": singular_label,
        "entity_label_plural": plural_label,
    }
    return RedirectResponse(url=f"{redirect_url}?report={report_id}", status_code=303)

CLIENT_FILTER_DEFINITIONS = [
    {
        "name": "status",
        "label": "Statut",
        "placeholder": "Tous les statuts",
        "options": list(STATUS_OPTIONS.items()),
    },
    {
        "name": "depannage",
        "label": "Dépannage",
        "placeholder": "Tous les dépannages",
        "options": list(DEPANNAGE_OPTIONS.items()),
    },
    {
        "name": "astreinte",
        "label": "Astreinte",
        "placeholder": "Toutes les astreintes",
        "options": [
            (key, label)
            for key, label in ASTREINTE_OPTIONS.items()
            if key != "pas_d_astreinte"
        ]
        + [("pas_d_astreinte", ASTREINTE_OPTIONS["pas_d_astreinte"])],
    },
]

FREQUENCY_UNITS = {
    "months": {
        "select_label": "mois",
        "singular_label": "mois",
        "plural_label": "mois",
        "use_plural_for_one": False,
    },
    "years": {
        "select_label": "années",
        "singular_label": "an",
        "plural_label": "ans",
        "use_plural_for_one": True,
    },
}

FREQUENCY_UNIT_OPTIONS = [(key, data["select_label"]) for key, data in FREQUENCY_UNITS.items()]

PREDEFINED_FREQUENCIES = {
    "contrat_annuel": {
        "label": "Contrat de maintenance annuel",
        "interval": 1,
        "unit": "years",
    },
    "contrat_semestriel": {
        "label": "Contrat de maintenance semestriel",
        "interval": 6,
        "unit": "months",
    },
    "contrat_trimestriel": {
        "label": "Contrat de maintenance trimestriel",
        "interval": 3,
        "unit": "months",
    },
    "contrat_bimestriel": {
        "label": "Contrat de maintenance bimestriel",
        "interval": 2,
        "unit": "months",
    },
    "contrat_mensuel": {
        "label": "Contrat de maintenance mensuel",
        "interval": 1,
        "unit": "months",
    },
    "prestation_ponctuelle": {
        "label": "Prestation ponctuelle",
        "interval": None,
        "unit": None,
    },
}

CUSTOM_INTERVAL_VALUE = "custom_interval"
CUSTOM_INTERVAL_LABEL = "Fréquence personnalisée"
INTERVAL_PREFIX = "interval:"

FREQUENCY_SELECT_OPTIONS = {
    key: data["label"] for key, data in PREDEFINED_FREQUENCIES.items()
}
FREQUENCY_SELECT_OPTIONS[CUSTOM_INTERVAL_VALUE] = CUSTOM_INTERVAL_LABEL

PREDEFINED_FREQUENCY_KEYS = tuple(PREDEFINED_FREQUENCIES.keys())

SUBCONTRACTING_FILTER_BASE = [
    {
        "name": "category",
        "label": "Catégorie",
        "placeholder": "Toutes les catégories",
    },
    {
        "name": "frequency",
        "label": "Fréquence",
        "placeholder": "Toutes les fréquences",
        "options": [(key, data["label"]) for key, data in PREDEFINED_FREQUENCIES.items()],
    },
]

def _slugify_identifier(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", value.lower())
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "prestation"


def _build_groups_from_definitions(
    definitions: Iterable[Any],
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, str]]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for definition in definitions:
        category = getattr(definition, "category", "Autres") or "Autres"
        group = grouped.setdefault(
            category,
            {
                "key": _slugify_identifier(category),
                "title": category,
                "options": [],
            },
        )
        option = {
            "value": getattr(definition, "key"),
            "label": getattr(definition, "label"),
            "budget_code": getattr(definition, "budget_code"),
            "position": getattr(definition, "position", 0) or 0,
        }
        group["options"].append(option)

    groups = sorted(
        grouped.values(),
        key=lambda item: (
            min((opt.get("position", 0) for opt in item["options"]), default=0),
            item["title"].lower(),
        ),
    )

    lookup: Dict[str, Dict[str, str]] = {}
    for group in groups:
        group["options"].sort(
            key=lambda opt: (opt.get("position", 0), opt["label"].lower())
        )
        for option in group["options"]:
            lookup[option["value"]] = {
                "label": option["label"],
                "budget_code": option["budget_code"],
                "category": group["title"],
            }
        for option in group["options"]:
            option.pop("position", None)
    return groups, lookup


def _build_groups_from_defaults() -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, str]]]:
    definitions = [
        SimpleNamespace(
            key=option["value"],
            label=option["label"],
            budget_code=option["budget_code"],
            category=group["category"],
            position=option.get("position", 0),
        )
        for group in DEFAULT_PRESTATION_GROUPS
        for option in group.get("options", [])
    ]
    return _build_groups_from_definitions(definitions)


def _get_subcontracted_options(
    session: Session,
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, str]]]:
    definitions = crud.list_prestation_definitions(session)
    if definitions:
        return _build_groups_from_definitions(definitions)
    return _build_groups_from_defaults()


def _build_category_filter_options(
    groups: Iterable[Dict[str, Any]]
) -> List[Tuple[str, str]]:
    seen: set[str] = set()
    options: List[Tuple[str, str]] = []
    for group in groups:
        title = group.get("title")
        if title and title not in seen:
            options.append((title, title))
            seen.add(title)
    return options


def _parse_interval_frequency(value: str) -> tuple[Optional[int], Optional[str]]:
    if not value.startswith(INTERVAL_PREFIX):
        return None, None
    parts = value.split(":")
    if len(parts) != 3:
        return None, None
    _, unit, interval_str = parts
    try:
        return int(interval_str), unit
    except ValueError:
        return None, None


def _format_interval_label(interval: int, unit: str) -> str:
    info = FREQUENCY_UNITS.get(unit)
    if not info:
        return f"Tous les {interval} {unit}"
    if interval == 1 and info.get("use_plural_for_one"):
        unit_label = info["plural_label"]
    elif interval == 1:
        unit_label = info["singular_label"]
    else:
        unit_label = info["plural_label"]
    if interval == 1:
        return f"Tous les {unit_label}"
    return f"Tous les {interval} {unit_label}"


def _frequency_label_from_details(
    frequency_value: str,
    interval: Optional[int],
    unit: Optional[str],
) -> str:
    predefined = PREDEFINED_FREQUENCIES.get(frequency_value)
    if predefined:
        return predefined["label"]
    resolved_interval = interval
    resolved_unit = unit
    if resolved_interval is None or not resolved_unit:
        parsed_interval, parsed_unit = _parse_interval_frequency(frequency_value)
        if resolved_interval is None:
            resolved_interval = parsed_interval
        if not resolved_unit:
            resolved_unit = parsed_unit
    if resolved_interval and resolved_unit in FREQUENCY_UNITS:
        return _format_interval_label(resolved_interval, resolved_unit)
    return frequency_value


def _resolve_frequency(
    selection: str,
    custom_interval: Optional[str],
    custom_unit: Optional[str],
) -> tuple[str, Optional[int], Optional[str]]:
    if selection == CUSTOM_INTERVAL_VALUE:
        if not custom_interval:
            raise HTTPException(400, "Veuillez renseigner l'intervalle de la fréquence")
        if custom_unit not in FREQUENCY_UNITS:
            raise HTTPException(400, "Unité de fréquence inconnue")
        try:
            interval_value = int(custom_interval)
        except ValueError as exc:
            raise HTTPException(400, "L'intervalle doit être un nombre entier") from exc
        if interval_value < 1:
            raise HTTPException(400, "L'intervalle doit être supérieur ou égal à 1")
        if interval_value > 120:
            raise HTTPException(400, "L'intervalle ne peut pas dépasser 120")
        frequency_value = f"{INTERVAL_PREFIX}{custom_unit}:{interval_value}"
        return frequency_value, interval_value, custom_unit
    if selection not in PREDEFINED_FREQUENCIES:
        raise HTTPException(400, "Fréquence inconnue")
    predefined = PREDEFINED_FREQUENCIES[selection]
    return selection, predefined["interval"], predefined["unit"]


def _build_frequency_labels(
    services: Iterable = (), extra_values: Iterable[str] = ()
) -> Dict[str, str]:
    labels = {key: data["label"] for key, data in PREDEFINED_FREQUENCIES.items()}
    for service in services:
        value = getattr(service, "frequency", None)
        if not value:
            continue
        interval = getattr(service, "frequency_interval", None)
        unit = getattr(service, "frequency_unit", None)
        labels[value] = _frequency_label_from_details(value, interval, unit)
    for value in extra_values:
        if not value or value == CUSTOM_INTERVAL_VALUE:
            continue
        if value not in labels:
            labels[value] = _frequency_label_from_details(value, None, None)
    return labels


def _build_frequency_filter_options(
    services: Iterable = (), extra_values: Iterable[str] = ()
) -> List[tuple[str, str]]:
    labels = _build_frequency_labels(services, extra_values)
    options: List[tuple[str, str]] = [
        (key, labels[key]) for key in PREDEFINED_FREQUENCY_KEYS if key in labels
    ]
    seen = {key for key, _ in options}
    for value in sorted(labels.keys()):
        if value in seen or value == CUSTOM_INTERVAL_VALUE:
            continue
        options.append((value, labels[value]))
        seen.add(value)
    return options


def _parse_budget(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    normalized = value.strip().replace(" ", "").replace(",", ".")
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError as exc:
        raise HTTPException(400, "Budget invalide") from exc


def _normalize_budget_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace("€", "")
        normalized = normalized.replace("\xa0", "").replace(" ", "")
        normalized = normalized.replace(",", ".")
        if not normalized:
            return None
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _format_currency_value(amount: float) -> str:
    formatted = f"{amount:,.2f}"
    formatted = formatted.replace(",", "\u00a0")
    return formatted.replace(".", ",")


def _clients_context(
    request: Request,
    clients,
    q: Optional[str],
    filters: Dict[str, str],
    entreprises,
    subcontracted_groups,
):
    report = _consume_import_report(request)
    all_services = [
        service
        for client in clients
        for service in getattr(client, "subcontractings", [])
    ]
    frequency_labels = _build_frequency_labels(all_services)
    return {
        "request": request,
        "clients": clients,
        "q": q or "",
        "depannage_options": DEPANNAGE_OPTIONS,
        "astreinte_options": ASTREINTE_OPTIONS,
        "status_options": STATUS_OPTIONS,
        "entreprises": entreprises,
        "entreprise_name_options": sorted({e.nom for e in entreprises}),
        "status_key_from_bool": _status_key_from_bool,
        "subcontract_status_options": SUBCONTRACT_STATUS_OPTIONS,
        "subcontract_status_styles": SUBCONTRACT_STATUS_STYLES,
        "subcontract_status_default": SUBCONTRACT_STATUS_DEFAULT,
        "subcontracted_groups": subcontracted_groups,
        "frequency_options": frequency_labels,
        "frequency_select_options": FREQUENCY_SELECT_OPTIONS,
        "frequency_unit_options": FREQUENCY_UNIT_OPTIONS,
        "custom_interval_value": CUSTOM_INTERVAL_VALUE,
        "predefined_frequency_keys": PREDEFINED_FREQUENCY_KEYS,
        "import_report": report,
        "focus_id": request.query_params.get("focus"),
        "active_filters": filters,
        "filters_definition": CLIENT_FILTER_DEFINITIONS,
    }

def _subcontractings_context(
    request: Request,
    services,
    q: Optional[str],
    filters: Dict[str, str],
    subcontracted_groups,
):
    report = _consume_import_report(request)
    category_totals: Dict[str, int] = {}
    frequency_totals: Dict[str, int] = {}
    total_budget = 0.0
    service_budget_display: Dict[int, str] = {}
    client_ids = set()

    for service in services:
        category_totals[service.category] = category_totals.get(service.category, 0) + 1
        frequency_totals[service.frequency] = frequency_totals.get(service.frequency, 0) + 1
        normalized_budget = _normalize_budget_value(getattr(service, "budget", None))
        if normalized_budget is not None:
            total_budget += normalized_budget
            if getattr(service, "id", None) is not None:
                service_budget_display[service.id] = _format_currency_value(
                    normalized_budget
                )
        if service.client_id:
            client_ids.add(service.client_id)

    active_frequency = filters.get("frequency")
    frequency_labels = _build_frequency_labels(services, [active_frequency] if active_frequency else [])
    frequency_filter_options = _build_frequency_filter_options(services, [active_frequency] if active_frequency else [])
    category_filter_options = _build_category_filter_options(subcontracted_groups)
    filters_definition = [
        {**SUBCONTRACTING_FILTER_BASE[0], "options": category_filter_options},
        {**SUBCONTRACTING_FILTER_BASE[1], "options": frequency_filter_options},
    ]

    return {
        "request": request,
        "services": services,
        "q": q or "",
        "frequency_options": frequency_labels,
        "frequency_select_options": FREQUENCY_SELECT_OPTIONS,
        "frequency_unit_options": FREQUENCY_UNIT_OPTIONS,
        "custom_interval_value": CUSTOM_INTERVAL_VALUE,
        "predefined_frequency_keys": PREDEFINED_FREQUENCY_KEYS,
        "category_totals": category_totals,
        "frequency_totals": frequency_totals,
        "total_budget": total_budget,
        "total_budget_formatted": _format_currency_value(total_budget),
        "total_services": len(services),
        "distinct_clients": len(client_ids),
        "service_budget_display": service_budget_display,
        "active_filters": filters,
        "filters_definition": filters_definition,
        "focus_id": request.query_params.get("focus"),
        "subcontract_status_options": SUBCONTRACT_STATUS_OPTIONS,
        "subcontract_status_styles": SUBCONTRACT_STATUS_STYLES,
        "subcontract_status_default": SUBCONTRACT_STATUS_DEFAULT,
        "subcontracted_groups": subcontracted_groups,
        "import_report": report,
    }


def _empty_subcontracted_service(
    default_client_id: Optional[int] = None,
) -> SimpleNamespace:
    default_frequency = (
        PREDEFINED_FREQUENCY_KEYS[0]
        if PREDEFINED_FREQUENCY_KEYS
        else CUSTOM_INTERVAL_VALUE
    )
    frequency_details = PREDEFINED_FREQUENCIES.get(default_frequency, {})
    return SimpleNamespace(
        id=None,
        prestation_key=None,
        prestation_label="Nouvelle prestation",
        budget_code="",
        budget=None,
        frequency=default_frequency,
        frequency_interval=frequency_details.get("interval"),
        frequency_unit=frequency_details.get("unit"),
        status=SUBCONTRACT_STATUS_DEFAULT,
        realization_week=None,
        order_week=None,
        client_id=default_client_id,
        client=None,
    )


def _extract_client_filters(
    status: Optional[str], depannage: Optional[str], astreinte: Optional[str]
) -> Dict[str, str]:
    filters: Dict[str, str] = {}
    if status in STATUS_OPTIONS:
        filters["status"] = status
    if depannage in DEPANNAGE_OPTIONS:
        filters["depannage"] = depannage
    if astreinte in ASTREINTE_OPTIONS:
        filters["astreinte"] = astreinte
    return filters


def _extract_subcontracting_filters(
    category: Optional[str],
    frequency: Optional[str],
    *,
    valid_categories: Optional[Iterable[str]] = None,
) -> Dict[str, str]:
    filters: Dict[str, str] = {}
    categories = set(valid_categories or [])
    if category and (not categories or category in categories):
        filters["category"] = category
    if frequency and frequency != CUSTOM_INTERVAL_VALUE:
        filters["frequency"] = frequency
    return filters


def _admin_prestations_context(
    request: Request,
    session: Session,
    *,
    errors: Optional[List[str]] = None,
    success: bool = False,
    updated: bool = False,
    focus: Optional[str] = None,
    form_values: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    definitions = crud.list_prestation_definitions(session)
    usage_counts = crud.count_subcontracted_services_by_definition(session)

    grouped: Dict[str, Dict[str, Any]] = {}
    sorted_definitions = sorted(
        definitions,
        key=lambda d: (
            (d.position or 0),
            (d.category or "").lower(),
            d.label.lower(),
        ),
    )

    for definition in sorted_definitions:
        category = definition.category or "Autres"
        group = grouped.setdefault(
            category,
            {
                "category": category,
                "slug": _slugify_identifier(category),
                "definitions": [],
            },
        )
        group["definitions"].append(definition)

    grouped_definitions = list(grouped.values())
    grouped_definitions.sort(
        key=lambda group: (
            min((d.position or 0) for d in group["definitions"]),
            group["category"].lower(),
        )
    )

    if not grouped_definitions and not definitions:
        # Provide a fallback view based on the defaults when no definition exists yet.
        default_groups, _ = _build_groups_from_defaults()
        grouped_definitions = [
            {
                "category": group["title"],
                "slug": group["key"],
                "definitions": [],
            }
            for group in default_groups
        ]

    form_defaults = {
        "label": "",
        "budget_code": "",
        "category": grouped_definitions[0]["category"] if grouped_definitions else "",
        "position": "0",
        "identifier": "",
    }
    if form_values:
        form_defaults.update(form_values)

    category_suggestions = sorted({definition.category for definition in definitions if definition.category})
    if not category_suggestions:
        category_suggestions = [group["category"] for group in DEFAULT_PRESTATION_GROUPS]

    return {
        "request": request,
        "grouped_definitions": grouped_definitions,
        "usage_counts": usage_counts,
        "errors": errors or None,
        "success": success,
        "updated": updated,
        "focus_key": focus,
        "form_values": form_defaults,
        "category_suggestions": category_suggestions,
    }


@app.on_event("startup")
def on_startup():
    init_db()
    _ensure_default_admin_user()
    app.state.import_reports = {}

# Page liste

@app.post("/token", response_model=Token)
def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    session: Session = Depends(get_session),
) -> Token:
    user = authenticate_user(session, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Identifiants invalides",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        {"sub": user.username}, expires_delta=access_token_expires
    )
    return Token(access_token=access_token, token_type="bearer")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: Optional[str] = None):
    cookie_token = request.cookies.get(SESSION_COOKIE_NAME)
    if cookie_token:
        try:
            payload = jwt.decode(cookie_token, SECRET_KEY, algorithms=[ALGORITHM])
            username = payload.get("sub")
            if username:
                with Session(engine) as session:
                    user = crud.get_user_by_username(session, username)
                if user:
                    request.state.user = user
                    return templates.TemplateResponse(
                        "login.html",
                        {
                            "request": request,
                            "next": next or "",
                            "already_authenticated": True,
                            "current_user": user,
                        },
                    )
        except JWTError:
            pass
    request.state.user = None
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "next": next or "",
            "current_user": None,
        },
    )


@app.post("/login")
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    user = authenticate_user(session, username, password)
    if not user:
        request.state.user = None
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Identifiants invalides",
                "username": username,
                "next": next or "",
                "current_user": None,
            },
            status_code=status.HTTP_401_UNAUTHORIZED,
        )
    user = crud.record_user_login(session, user)
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        {"sub": user.username}, expires_delta=access_token_expires
    )
    redirect_target = next or "/"
    response = RedirectResponse(
        url=redirect_target,
        status_code=status.HTTP_303_SEE_OTHER,
    )
    response.set_cookie(
        SESSION_COOKIE_NAME,
        access_token,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        httponly=True,
        secure=SESSION_COOKIE_SECURE,
        samesite="lax",
    )
    return response


@app.post("/logout")
def logout(
    _current_user: CurrentUser,
    redirect_to: Optional[str] = Form("/login"),
    session: Session = Depends(get_session),
):
    db_user = crud.get_user_by_username(session, _current_user.username)
    if db_user:
        crud.record_user_logout(session, db_user)
    response = RedirectResponse(
        url=redirect_to or "/login",
        status_code=status.HTTP_303_SEE_OTHER,
    )
    response.delete_cookie(SESSION_COOKIE_NAME)
    return response


@app.get("/mon-compte", response_class=HTMLResponse)
def account_page(request: Request, current_user: CurrentUser):
    success = request.query_params.get("success") == "1"
    return templates.TemplateResponse(
        "account.html",
        {
            "request": request,
            "current_user": current_user,
            "errors": [],
            "success": success,
            "password_min_length": PASSWORD_MIN_LENGTH,
        },
    )


@app.post("/mon-compte/mot-de-passe", response_class=HTMLResponse)
def account_change_password(
    request: Request,
    current_user: CurrentUser,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    session: Session = Depends(get_session),
):
    errors: List[str] = []
    if not verify_password(current_password, current_user.hashed_password):
        errors.append("Le mot de passe actuel est incorrect.")
    if len(new_password) < PASSWORD_MIN_LENGTH:
        errors.append(
            f"Le nouveau mot de passe doit contenir au moins {PASSWORD_MIN_LENGTH} caractères."
        )
    if new_password != confirm_password:
        errors.append("La confirmation du mot de passe ne correspond pas.")

    db_user = crud.get_user_by_username(session, current_user.username)
    if not db_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Utilisateur introuvable.",
        )

    if not errors:
        try:
            hashed_password = get_password_hash(new_password)
        except ValueError as exc:
            errors.append(str(exc))
        else:
            updated_user = crud.update_user_password(session, db_user, hashed_password)
            current_user.hashed_password = updated_user.hashed_password
            request.state.user = updated_user
            response = RedirectResponse(
                url="/mon-compte?success=1",
                status_code=status.HTTP_303_SEE_OTHER,
            )
            return response

    request.state.user = current_user
    return templates.TemplateResponse(
        "account.html",
        {
            "request": request,
            "current_user": current_user,
            "errors": errors,
            "success": False,
            "password_min_length": PASSWORD_MIN_LENGTH,
        },
        status_code=status.HTTP_400_BAD_REQUEST,
    )


@app.get("/admin/utilisateurs", response_class=HTMLResponse)
def admin_users_page(
    request: Request,
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    users = crud.list_users(session)
    login_history = crud.get_login_history_for_users(
        session, [user.id for user in users if user.id], limit=5
    )
    return templates.TemplateResponse(
        "admin_users.html",
        {
            "request": request,
            "users": users,
            "errors": None,
            "success": request.query_params.get("success") == "1",
            "form_values": {"username": ""},
            "focus_username": request.query_params.get("focus"),
            "login_history": login_history,
            "reset_result": None,
        },
    )


@app.post("/admin/utilisateurs", response_class=HTMLResponse)
def admin_users_create(
    request: Request,
    _current_user: CurrentUser,
    username: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    trimmed_username = username.strip()
    errors: List[str] = []
    if not trimmed_username:
        errors.append("L'identifiant est obligatoire.")
    if " " in trimmed_username:
        errors.append("L'identifiant ne peut pas contenir d'espaces.")
    if not password:
        errors.append("Le mot de passe est obligatoire.")
    if trimmed_username and crud.get_user_by_username(session, trimmed_username):
        errors.append("Cet identifiant est déjà utilisé.")

    hashed_password: Optional[str] = None
    if not errors:
        try:
            hashed_password = get_password_hash(password)
        except ValueError as exc:
            errors.append(str(exc))

    if errors or hashed_password is None:
        users = crud.list_users(session)
        login_history = crud.get_login_history_for_users(
            session, [user.id for user in users if user.id], limit=5
        )
        return templates.TemplateResponse(
            "admin_users.html",
            {
                "request": request,
                "users": users,
                "errors": errors,
                "success": False,
                "form_values": {"username": trimmed_username},
                "focus_username": None,
                "login_history": login_history,
                "reset_result": None,
            },
            status_code=status.HTTP_400_BAD_REQUEST,
        )

    crud.create_user(
        session,
        UserCreate(username=trimmed_username, hashed_password=hashed_password),
    )
    return RedirectResponse(
        url=f"/admin/utilisateurs?success=1&focus={quote(trimmed_username)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@app.post(
    "/admin/utilisateurs/{username}/reinitialiser-mot-de-passe",
    response_class=HTMLResponse,
)
def admin_users_reset_password(
    request: Request,
    _current_user: CurrentUser,
    username: str,
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    errors: List[str] = []
    reset_result: Optional[Dict[str, str]] = None
    target_user = crud.get_user_by_username(session, username)
    if not target_user:
        errors.append("Utilisateur introuvable.")
    else:
        try:
            temporary_password = _generate_temporary_password()
            hashed_password = get_password_hash(temporary_password)
        except ValueError as exc:
            errors.append(str(exc))
        else:
            crud.update_user_password(session, target_user, hashed_password)
            reset_result = {
                "username": target_user.username,
                "password": temporary_password,
            }
    users = crud.list_users(session)
    login_history = crud.get_login_history_for_users(
        session, [user.id for user in users if user.id], limit=5
    )
    status_code = status.HTTP_400_BAD_REQUEST if errors else status.HTTP_200_OK
    return templates.TemplateResponse(
        "admin_users.html",
        {
            "request": request,
            "users": users,
            "errors": errors or None,
            "success": False,
            "form_values": {"username": ""},
            "focus_username": reset_result["username"] if reset_result else None,
            "login_history": login_history,
            "reset_result": reset_result,
        },
        status_code=status_code,
    )


@app.get("/admin/prestations", response_class=HTMLResponse)
def admin_prestations_page(
    request: Request,
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    context = _admin_prestations_context(
        request,
        session,
        success=request.query_params.get("success") == "1",
        updated=request.query_params.get("updated") == "1",
        focus=request.query_params.get("focus"),
    )
    return templates.TemplateResponse("admin_prestations.html", context)


@app.post("/admin/prestations", response_class=HTMLResponse)
def admin_prestations_create(
    request: Request,
    _current_user: CurrentUser,
    label: str = Form(...),
    budget_code: str = Form(...),
    category: str = Form(...),
    position: Optional[str] = Form("0"),
    identifier: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    trimmed_label = label.strip()
    trimmed_category = category.strip()
    trimmed_budget_code = budget_code.strip()
    slug_source = (identifier or trimmed_label).strip()
    generated_key = _slugify_identifier(slug_source)
    errors: List[str] = []
    if not trimmed_label:
        errors.append("Le libellé est obligatoire.")
    if not trimmed_category:
        errors.append("La catégorie est obligatoire.")
    if not trimmed_budget_code:
        errors.append("Le code budget est obligatoire.")
    if not generated_key:
        errors.append("L'identifiant doit contenir au moins un caractère alphanumérique.")
    if generated_key and crud.get_prestation_definition_by_key(session, generated_key):
        errors.append("Cet identifiant est déjà utilisé.")

    try:
        position_value = int(position or 0)
        if position_value < 0:
            raise ValueError
    except ValueError:
        errors.append("La position doit être un entier positif.")
        position_value = 0

    if errors:
        context = _admin_prestations_context(
            request,
            session,
            errors=errors,
            form_values={
                "label": trimmed_label,
                "budget_code": trimmed_budget_code,
                "category": trimmed_category,
                "position": str(position_value),
                "identifier": slug_source,
            },
        )
        return templates.TemplateResponse(
            "admin_prestations.html",
            context,
            status_code=status.HTTP_400_BAD_REQUEST,
        )

    definition = crud.create_prestation_definition(
        session,
        PrestationDefinitionCreate(
            key=generated_key,
            label=trimmed_label,
            budget_code=trimmed_budget_code,
            category=trimmed_category,
            position=position_value,
        ),
    )
    return RedirectResponse(
        url=f"/admin/prestations?success=1&focus={quote(definition.key)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@app.post("/admin/prestations/{definition_id}/update", response_class=HTMLResponse)
def admin_prestations_update(
    request: Request,
    _current_user: CurrentUser,
    definition_id: int,
    label: str = Form(...),
    budget_code: str = Form(...),
    category: str = Form(...),
    position: Optional[str] = Form("0"),
    session: Session = Depends(get_session),
):
    _ensure_admin_access(_current_user)
    definition = crud.get_prestation_definition(session, definition_id)
    if not definition:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Prestation introuvable")

    trimmed_label = label.strip()
    trimmed_category = category.strip()
    trimmed_budget_code = budget_code.strip()
    errors: List[str] = []

    if not trimmed_label:
        errors.append("Le libellé est obligatoire.")
    if not trimmed_category:
        errors.append("La catégorie est obligatoire.")
    if not trimmed_budget_code:
        errors.append("Le code budget est obligatoire.")

    try:
        position_value = int(position or 0)
        if position_value < 0:
            raise ValueError
    except ValueError:
        errors.append("La position doit être un entier positif.")
        position_value = definition.position or 0

    if errors:
        context = _admin_prestations_context(
            request,
            session,
            errors=errors,
            form_values={
                "label": trimmed_label,
                "budget_code": trimmed_budget_code,
                "category": trimmed_category,
                "position": str(position_value),
                "identifier": definition.key,
            },
            focus=definition.key,
        )
        return templates.TemplateResponse(
            "admin_prestations.html",
            context,
            status_code=status.HTTP_400_BAD_REQUEST,
        )

    updated_definition = crud.update_prestation_definition(
        session,
        definition_id,
        PrestationDefinitionUpdate(
            label=trimmed_label,
            budget_code=trimmed_budget_code,
            category=trimmed_category,
            position=position_value,
        ),
    )
    if not updated_definition:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Prestation introuvable")

    crud.sync_subcontracted_services_from_definition(session, updated_definition)

    return RedirectResponse(
        url=f"/admin/prestations?updated=1&focus={quote(updated_definition.key)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


# Page liste
@app.get("/", response_class=HTMLResponse)
def clients_page(
    request: Request,
    _current_user: CurrentUser,
    q: Optional[str] = None,
    status: Optional[str] = None,
    depannage: Optional[str] = None,
    astreinte: Optional[str] = None,
    session: Session = Depends(get_session),
): 
    filters = _extract_client_filters(status, depannage, astreinte)
    clients = crud.list_clients(session, q=q, filters=filters)
    entreprises = crud.list_entreprises(session)
    subcontracted_groups, _ = _get_subcontracted_options(session)
    return templates.TemplateResponse(
        "clients_list.html",
        _clients_context(
            request,
            clients,
            q,
            filters,
            entreprises,
            subcontracted_groups,
        ),
    )

# Fragment liste (HTMX)
@app.get("/_clients", response_class=HTMLResponse)
def clients_fragment(
    request: Request,
    _current_user: CurrentUser,
    q: Optional[str] = None,
    status: Optional[str] = None,
    depannage: Optional[str] = None,
    astreinte: Optional[str] = None,
    session: Session = Depends(get_session),
): 
    filters = _extract_client_filters(status, depannage, astreinte)
    clients = crud.list_clients(session, q=q, filters=filters)
    entreprises = crud.list_entreprises(session)
    subcontracted_groups, _ = _get_subcontracted_options(session)
    return templates.TemplateResponse(
        "clients_list.html",
        _clients_context(
            request,
            clients,
            q,
            filters,
            entreprises,
            subcontracted_groups,
        ),
    )


@app.get("/prestations", response_class=HTMLResponse)
def subcontracted_services_page(
    request: Request,
    _current_user: CurrentUser,
    q: Optional[str] = None,
    category: Optional[str] = None,
    frequency: Optional[str] = None,
    session: Session = Depends(get_session),
):
    subcontracted_groups, _ = _get_subcontracted_options(session)
    valid_categories = [group.get("title") for group in subcontracted_groups]
    filters = _extract_subcontracting_filters(
        category,
        frequency,
        valid_categories=[c for c in valid_categories if c],
    )
    services = crud.list_subcontracted_services(session, q=q, filters=filters)
    return templates.TemplateResponse(
        "subcontractings_list.html",
        _subcontractings_context(
            request,
            services,
            q,
            filters,
            subcontracted_groups,
        ),
    )


@app.get("/prestations/{service_id}/edit", response_class=HTMLResponse)
def subcontracted_service_edit_page(
    request: Request,
    _current_user: CurrentUser,
    service_id: int,
    session: Session = Depends(get_session),
):
    service = crud.get_subcontracted_service(session, service_id)
    if not service:
        raise HTTPException(404, "Prestation introuvable")

    subcontracted_groups, subcontracted_lookup = _get_subcontracted_options(session)
    available_keys = set(subcontracted_lookup.keys())
    clients = crud.list_client_choices(session)
    return templates.TemplateResponse(
        "subcontracting_edit.html",
        {
            "request": request,
            "service": service,
            "subcontracted_groups": subcontracted_groups,
            "frequency_options": _build_frequency_labels([service]),
            "frequency_select_options": FREQUENCY_SELECT_OPTIONS,
            "frequency_unit_options": FREQUENCY_UNIT_OPTIONS,
            "custom_interval_value": CUSTOM_INTERVAL_VALUE,
            "predefined_frequency_keys": PREDEFINED_FREQUENCY_KEYS,
            "subcontract_status_options": SUBCONTRACT_STATUS_OPTIONS,
            "subcontract_status_default": SUBCONTRACT_STATUS_DEFAULT,
            "available_prestations": available_keys,
            "clients": clients,
            "return_url": f"/prestations?focus={service.id}#service-{service.id}",
            "form_action": f"/prestations/{service.id}/edit",
            "is_creation": False,
        },
    )


@app.get("/prestations/new", response_class=HTMLResponse)
def subcontracted_service_create_page(
    request: Request,
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
):
    subcontracted_groups, subcontracted_lookup = _get_subcontracted_options(session)
    clients = crud.list_client_choices(session)
    default_client_id = clients[0].id if clients else None
    empty_service = _empty_subcontracted_service(default_client_id)
    available_keys = set(subcontracted_lookup.keys())
    return templates.TemplateResponse(
        "subcontracting_edit.html",
        {
            "request": request,
            "service": empty_service,
            "subcontracted_groups": subcontracted_groups,
            "frequency_options": _build_frequency_labels([]),
            "frequency_select_options": FREQUENCY_SELECT_OPTIONS,
            "frequency_unit_options": FREQUENCY_UNIT_OPTIONS,
            "custom_interval_value": CUSTOM_INTERVAL_VALUE,
            "predefined_frequency_keys": PREDEFINED_FREQUENCY_KEYS,
            "subcontract_status_options": SUBCONTRACT_STATUS_OPTIONS,
            "subcontract_status_default": SUBCONTRACT_STATUS_DEFAULT,
            "available_prestations": available_keys,
            "clients": clients,
            "return_url": "/prestations",
            "form_action": "/prestations/new",
            "is_creation": True,
        },
    )


@app.post("/prestations/{service_id}/edit")
def update_subcontracted_service(
    _current_user: CurrentUser,
    service_id: int,
    prestation: str = Form(...),
    client_id: int = Form(...),
    budget: Optional[str] = Form(None),
    frequency: str = Form(...),
    custom_frequency_interval: Optional[str] = Form(None),
    custom_frequency_unit: Optional[str] = Form(None),
    status: str = Form(SUBCONTRACT_STATUS_DEFAULT),
    realization_week: Optional[str] = Form(None),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    service = crud.get_subcontracted_service(session, service_id)
    if not service:
        raise HTTPException(404, "Prestation introuvable")

    if not crud.get_client(session, client_id):
        raise HTTPException(400, "Client inconnu")

    (
        resolved_frequency,
        resolved_interval,
        resolved_unit,
    ) = _resolve_frequency(frequency, custom_frequency_interval, custom_frequency_unit)

    _, subcontracted_lookup = _get_subcontracted_options(session)
    details = subcontracted_lookup.get(prestation)
    if not details and prestation != service.prestation_key:
        raise HTTPException(400, "Prestation inconnue")

    if status not in SUBCONTRACT_STATUS_OPTIONS:
        raise HTTPException(400, "Statut inconnu")

    parsed_budget = _parse_budget(budget)

    realization_value: Optional[str] = None
    if resolved_frequency == "prestation_ponctuelle" and realization_week:
        realization_value = realization_week.strip().upper()

    order_value = order_week.strip().upper() if order_week else None

    update_payload = SubcontractedServiceUpdate(
        prestation_key=prestation,
        prestation_label=(
            details["label"] if details else service.prestation_label
        ),
        category=(details["category"] if details else service.category),
        budget_code=(details["budget_code"] if details else service.budget_code),
        budget=parsed_budget,
        frequency=resolved_frequency,
        frequency_interval=resolved_interval,
        frequency_unit=resolved_unit,
        status=status,
        realization_week=realization_value,
        order_week=order_value,
        client_id=client_id,
    )

    updated = crud.update_subcontracted_service(session, service_id, update_payload)
    if not updated:
        raise HTTPException(404, "Prestation introuvable")

    return RedirectResponse(
        url=f"/prestations?focus={service_id}#service-{service_id}",
        status_code=303,
    )


def _create_subcontracted_service_from_form(
    session: Session,
    *,
    client_id: int,
    prestation: str,
    budget: Optional[str],
    frequency: str,
    custom_frequency_interval: Optional[str],
    custom_frequency_unit: Optional[str],
    status: str,
    realization_week: Optional[str],
    order_week: Optional[str],
):
    _, subcontracted_lookup = _get_subcontracted_options(session)
    details = subcontracted_lookup.get(prestation)
    if not details:
        raise HTTPException(400, "Prestation inconnue")

    (
        resolved_frequency,
        resolved_interval,
        resolved_unit,
    ) = _resolve_frequency(
        frequency, custom_frequency_interval, custom_frequency_unit
    )

    if status not in SUBCONTRACT_STATUS_OPTIONS:
        raise HTTPException(400, "Statut inconnu")

    parsed_budget = _parse_budget(budget)
    realization_value = (
        realization_week.strip().upper()
        if realization_week and resolved_frequency == "prestation_ponctuelle"
        else None
    )
    order_value = order_week.strip().upper() if order_week else None

    created = crud.create_subcontracted_service(
        session,
        client_id,
        SubcontractedServiceCreate(
            prestation_key=prestation,
            prestation_label=details["label"],
            category=details["category"],
            budget_code=details["budget_code"],
            budget=parsed_budget,
            frequency=resolved_frequency,
            frequency_interval=resolved_interval,
            frequency_unit=resolved_unit,
            status=status,
            realization_week=realization_value,
            order_week=order_value,
        ),
    )
    if not created:
        raise HTTPException(404, "Client introuvable")
    return created


def _resolve_import_client_id(session: Session, row: Dict[str, Any]) -> int:
    raw_client_id = row.get("client_id")
    if raw_client_id:
        client = crud.get_client(session, int(raw_client_id))
        if not client:
            raise ValueError(f"Client #{raw_client_id} introuvable")
        return client.id

    company_name = (row.get("company_name") or "").strip()
    client_name = (row.get("client_name") or "").strip()

    matches = crud.find_clients_for_import(
        session,
        company_name=company_name or None,
        client_name=client_name or None,
    )

    if not matches:
        if company_name and client_name:
            raise ValueError(
                f"Aucun client trouvé pour {company_name} / {client_name}."
            )
        if company_name:
            raise ValueError(f"Aucun client trouvé pour {company_name}.")
        raise ValueError(
            "Aucun client correspondant au contact renseigné. Ajoutez l'identifiant client pour lever l'ambiguïté."
        )

    if len(matches) > 1:
        raise ValueError(
            "Plusieurs clients correspondent aux informations fournies. Précisez l'identifiant client pour lever l'ambiguïté."
        )

    return matches[0].id


def _resolve_import_prestation_key(
    row: Dict[str, Any], lookup: Dict[str, Dict[str, str]]
) -> str:
    raw_value = (row.get("prestation") or "").strip()
    label_value = (row.get("prestation_label") or "").strip()

    if raw_value:
        if raw_value in lookup:
            return raw_value
        normalized_value = _slugify_identifier(raw_value)
        if normalized_value in lookup:
            return normalized_value
        for key in lookup.keys():
            if _slugify_identifier(key) == normalized_value:
                return key

    if label_value:
        normalized_label = _slugify_identifier(label_value)
        for key, details in lookup.items():
            if _slugify_identifier(details.get("label", "")) == normalized_label:
                return key

    target = raw_value or label_value or "(libellé manquant)"
    raise ValueError(f"Prestation '{target}' introuvable dans le référentiel.")


def _prepare_import_frequency(row: Dict[str, Any]) -> Tuple[str, Optional[str], Optional[str]]:
    raw_frequency = (row.get("frequency") or "").strip()
    if not raw_frequency:
        raise ValueError("La fréquence est obligatoire.")

    if raw_frequency.startswith(INTERVAL_PREFIX):
        interval, unit = _parse_interval_frequency(raw_frequency)
        if not interval or not unit:
            raise ValueError("Fréquence personnalisée invalide.")
        return CUSTOM_INTERVAL_VALUE, str(interval), unit

    if raw_frequency == CUSTOM_INTERVAL_VALUE:
        interval_value = row.get("frequency_interval")
        unit_value = row.get("frequency_unit")
        if interval_value is None or not unit_value:
            raise ValueError(
                "Indiquez l'intervalle et l'unité pour la fréquence personnalisée."
            )
        return CUSTOM_INTERVAL_VALUE, str(interval_value), unit_value

    return raw_frequency, None, None


@app.post("/prestations/new")
def create_subcontracted_service_from_view(
    _current_user: CurrentUser,
    prestation: str = Form(...),
    client_id: int = Form(...),
    budget: Optional[str] = Form(None),
    frequency: str = Form(...),
    custom_frequency_interval: Optional[str] = Form(None),
    custom_frequency_unit: Optional[str] = Form(None),
    status: str = Form(SUBCONTRACT_STATUS_DEFAULT),
    realization_week: Optional[str] = Form(None),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    created = _create_subcontracted_service_from_form(
        session,
        client_id=client_id,
        prestation=prestation,
        budget=budget,
        frequency=frequency,
        custom_frequency_interval=custom_frequency_interval,
        custom_frequency_unit=custom_frequency_unit,
        status=status,
        realization_week=realization_week,
        order_week=order_week,
    )
    return RedirectResponse(
        url=f"/prestations?focus={created.id}#service-{created.id}",
        status_code=303,
    )


@app.post("/prestations/import")
async def import_subcontracted_services(
    _current_user: CurrentUser,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    if not file.filename:
        return _store_import_report(
            "/prestations",
            created=0,
            total=0,
            errors=["Aucun fichier sélectionné."],
            filename="Import prestations",
            singular_label="prestation",
            plural_label="prestations",
        )

    if not file.filename.lower().endswith(ALLOWED_IMPORT_EXTENSIONS):
        return _store_import_report(
            "/prestations",
            created=0,
            total=0,
            errors=[
                "Format de fichier non supporté. Merci d'utiliser un fichier Excel (.xlsx)."
            ],
            filename=file.filename,
            singular_label="prestation",
            plural_label="prestations",
        )

    content = await file.read()
    try:
        rows = parse_prestations_excel(content)
    except ValueError as exc:
        return _store_import_report(
            "/prestations",
            created=0,
            total=0,
            errors=[str(exc)],
            filename=file.filename,
            singular_label="prestation",
            plural_label="prestations",
        )

    _, subcontracted_lookup = _get_subcontracted_options(session)
    created = 0
    errors: List[str] = []

    for idx, row in enumerate(rows, start=1):
        row_number = row.get("__row__", idx)
        try:
            client_id = _resolve_import_client_id(session, row)
            prestation_key = _resolve_import_prestation_key(row, subcontracted_lookup)
            frequency_value, custom_interval, custom_unit = _prepare_import_frequency(row)
            status_value = row.get("status") or SUBCONTRACT_STATUS_DEFAULT
            if status_value not in SUBCONTRACT_STATUS_OPTIONS:
                raise ValueError(
                    f"Statut '{row.get('status')}' inconnu. Valeurs acceptées: "
                    + ", ".join(SUBCONTRACT_STATUS_OPTIONS.keys())
                )

            realization_week = row.get("realization_week")
            order_week = row.get("order_week")

            created_service = _create_subcontracted_service_from_form(
                session,
                client_id=client_id,
                prestation=prestation_key,
                budget=row.get("budget"),
                frequency=frequency_value,
                custom_frequency_interval=custom_interval,
                custom_frequency_unit=custom_unit,
                status=status_value,
                realization_week=(
                    realization_week.strip().upper()
                    if isinstance(realization_week, str)
                    else realization_week
                ),
                order_week=(
                    order_week.strip().upper()
                    if isinstance(order_week, str)
                    else order_week
                ),
            )
            if created_service:
                created += 1
        except (HTTPException, ValueError) as exc:
            session.rollback()
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            errors.append(f"Ligne {row_number} : {detail}")

    return _store_import_report(
        "/prestations",
        created=created,
        total=len(rows),
        errors=errors,
        filename=file.filename,
        singular_label="prestation",
        plural_label="prestations",
    )

# Form création
@app.post("/clients/new")
def create_client(
    _current_user: CurrentUser,
    company_name: str = Form(...),
    name: str = Form(...),
    email: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    technician_name: Optional[str] = Form(None),
    billing_address: Optional[str] = Form(None),
    depannage: str = Form("non_refacturable"),
    astreinte: str = Form("pas_d_astreinte"),
    tags: Optional[str] = Form(None),
    status: Optional[str] = Form("actif"),
    session: Session = Depends(get_session)
):
    contacts_payload: List[ContactCreate] = []
    if name or email or phone:
        contacts_payload.append(ContactCreate(name=name, email=email, phone=phone))

    statut_bool = _status_to_bool(status)
    entreprise = crud.ensure_entreprise(
        session,
        name=company_name,
        adresse_facturation=billing_address,
        tag=tags,
        statut=statut_bool,
    )
    client = crud.create_client(
        session,
        ClientCreate(
            entreprise_id=entreprise.id,
            name=name,
            email=email,
            phone=phone,
            technician_name=technician_name,
            billing_address=billing_address,
            depannage=depannage,
            astreinte=astreinte,
            tags=tags,
            status=status,
        ),
        contacts=contacts_payload,
    )
    return RedirectResponse(
        url=f"/?focus={client.id}#client-{client.id}", status_code=303
    )

# Maj
@app.post("/clients/{client_id}/edit")
def edit_client(
    _current_user: CurrentUser,
    client_id: int,
    company_name: str = Form(...),
    name: str = Form(...),
    email: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    technician_name: Optional[str] = Form(None),
    billing_address: Optional[str] = Form(None),
    depannage: str = Form("non_refacturable"),
    astreinte: str = Form("pas_d_astreinte"),
    tags: Optional[str] = Form(None),
    status: Optional[str] = Form("actif"),
    session: Session = Depends(get_session)
):
    statut_bool = _status_to_bool(status)
    entreprise = crud.ensure_entreprise(
        session,
        name=company_name,
        adresse_facturation=billing_address,
        tag=tags,
        statut=statut_bool,
    )
    updated = crud.update_client(
        session,
        client_id,
        ClientUpdate(
            entreprise_id=entreprise.id,
            name=name,
            email=email,
            phone=phone,
            technician_name=technician_name,
            billing_address=billing_address,
            depannage=depannage,
            astreinte=astreinte,
            tags=tags,
            status=status,
        ),
    )
    if not updated: raise HTTPException(404, "Client introuvable")
    return RedirectResponse(url="/", status_code=303)

@app.post("/clients/{client_id}/delete")
def remove_client(
    _current_user: CurrentUser,
    client_id: int,
    session: Session = Depends(get_session),
):
    ok = crud.delete_client(session, client_id)
    if not ok: raise HTTPException(404, "Client introuvable")
    return RedirectResponse(url="/", status_code=303)


@app.post("/clients/{client_id}/contacts")
def add_contact(
    _current_user: CurrentUser,
    client_id: int,
    name: str = Form(..., alias="contact_name"),
    email: Optional[str] = Form(None, alias="contact_email"),
    phone: Optional[str] = Form(None, alias="contact_phone"),
    session: Session = Depends(get_session),
):
    created = crud.create_contact(
        session,
        client_id,
        ContactCreate(name=name, email=email, phone=phone),
    )
    if not created:
        raise HTTPException(404, "Client introuvable")
    return RedirectResponse(
        url=f"/?focus={client_id}#client-{client_id}", status_code=303
    )


@app.post("/clients/{client_id}/contacts/{contact_id}/delete")
def remove_contact(
    _current_user: CurrentUser,
    client_id: int,
    contact_id: int,
    session: Session = Depends(get_session),
):
    ok = crud.delete_contact(session, client_id, contact_id)
    if not ok:
        raise HTTPException(404, "Contact introuvable")
    return RedirectResponse(
        url=f"/?focus={client_id}#client-{client_id}", status_code=303
    )


@app.post("/clients/{client_id}/subcontractings")
def add_subcontracted_service(
    _current_user: CurrentUser,
    client_id: int,
    prestation: str = Form(...),
    budget: Optional[str] = Form(None),
    frequency: str = Form(...),
    custom_frequency_interval: Optional[str] = Form(None),
    custom_frequency_unit: Optional[str] = Form(None),
    status: str = Form(SUBCONTRACT_STATUS_DEFAULT),
    realization_week: Optional[str] = Form(None),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    created = _create_subcontracted_service_from_form(
        session,
        client_id=client_id,
        prestation=prestation,
        budget=budget,
        frequency=frequency,
        custom_frequency_interval=custom_frequency_interval,
        custom_frequency_unit=custom_frequency_unit,
        status=status,
        realization_week=realization_week,
        order_week=order_week,
    )
    return RedirectResponse(
        url=f"/?focus={client_id}#client-{client_id}", status_code=303
    )


@app.post("/clients/{client_id}/subcontractings/{service_id}/delete")
def remove_subcontracted_service(
    _current_user: CurrentUser,
    client_id: int,
    service_id: int,
    session: Session = Depends(get_session),
):
    ok = crud.delete_subcontracted_service(session, client_id, service_id)
    if not ok:
        raise HTTPException(404, "Prestation introuvable")
    return RedirectResponse(
        url=f"/?focus={client_id}#client-{client_id}", status_code=303
    )


@app.get("/filtres-courroies", response_class=HTMLResponse)
def list_filters_and_belts(
    request: Request,
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
):
    filters_q = (request.query_params.get("filters_q") or "").strip()
    belts_q = (request.query_params.get("belts_q") or "").strip()

    filters = crud.list_filter_lines(session, q=filters_q or None)
    belts = crud.list_belt_lines(session, q=belts_q or None)
    editing_filter = None
    editing_belt = None

    edit_filter_id = request.query_params.get("edit_filter")
    if edit_filter_id:
        try:
            editing_filter = crud.get_filter_line(session, int(edit_filter_id))
        except ValueError:
            raise HTTPException(status_code=404, detail="Ligne filtre introuvable")
        if not editing_filter:
            raise HTTPException(status_code=404, detail="Ligne filtre introuvable")

    edit_belt_id = request.query_params.get("edit_belt")
    if edit_belt_id:
        try:
            editing_belt = crud.get_belt_line(session, int(edit_belt_id))
        except ValueError:
            raise HTTPException(status_code=404, detail="Ligne courroie introuvable")
        if not editing_belt:
            raise HTTPException(status_code=404, detail="Ligne courroie introuvable")

    return templates.TemplateResponse(
        "filters_belts.html",
        {
            "request": request,
            "filters": filters,
            "belts": belts,
            "filter_format_options": FILTER_FORMAT_OPTIONS,
            "filter_format_labels": FILTER_FORMAT_LABELS,
            "import_report": _consume_import_report(request),
            "editing_filter": editing_filter,
            "editing_belt": editing_belt,
            "filters_q": filters_q,
            "belts_q": belts_q,
        },
    )


@app.get("/plan-de-charge", response_class=HTMLResponse)
def workload_plan(
    request: Request,
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
) -> HTMLResponse:
    site_count = len(crud.list_workload_sites(session))
    return templates.TemplateResponse(
        "plan_de_charge.html",
        {
            "request": request,
            "workload_site_count": site_count,
        },
    )


class WorkloadPlanSitePayload(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)


class WorkloadPlanCellPayload(BaseModel):
    site_id: int
    day_index: int
    value: Optional[str] = None


class WorkloadPlanCellsPayload(BaseModel):
    updates: List[WorkloadPlanCellPayload] = Field(default_factory=list)


class WorkloadPlanImportPayload(BaseModel):
    version: Optional[int] = None
    sites: List[str] = Field(default_factory=list)
    cells: Dict[str, List[Optional[str]]] = Field(default_factory=dict)


class WorkloadPlanSiteResponse(BaseModel):
    id: int
    name: str
    position: int
    cells: List[str]


class WorkloadPlanResponse(BaseModel):
    version: int
    sites: List[WorkloadPlanSiteResponse]


@app.get("/api/workload-plan", response_model=WorkloadPlanResponse)
def get_workload_plan(
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
) -> WorkloadPlanResponse:
    sites = crud.list_workload_sites(session)
    payload_sites: List[WorkloadPlanSiteResponse] = []
    for site in sites:
        cells = ["" for _ in range(364)]
        for cell in site.cells:
            if 0 <= cell.day_index < 364 and cell.value:
                cells[cell.day_index] = cell.value
        payload_sites.append(
            WorkloadPlanSiteResponse(
                id=site.id,
                name=site.name,
                position=site.position,
                cells=cells,
            )
        )
    return WorkloadPlanResponse(version=1, sites=payload_sites)


@app.post(
    "/api/workload-plan/sites",
    response_model=WorkloadPlanSiteResponse,
    status_code=201,
)
def create_workload_plan_site(
    _current_user: CurrentUser,
    payload: WorkloadPlanSitePayload,
    session: Session = Depends(get_session),
) -> WorkloadPlanSiteResponse:
    try:
        site = crud.create_workload_site(
            session, WorkloadSiteCreate(name=payload.name)
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return WorkloadPlanSiteResponse(
        id=site.id,
        name=site.name,
        position=site.position,
        cells=["" for _ in range(364)],
    )


@app.patch("/api/workload-plan/sites/{site_id}", response_model=WorkloadPlanSiteResponse)
def rename_workload_plan_site(
    _current_user: CurrentUser,
    site_id: int,
    payload: WorkloadPlanSitePayload,
    session: Session = Depends(get_session),
) -> WorkloadPlanSiteResponse:
    try:
        site = crud.rename_workload_site(
            session, site_id, WorkloadSiteUpdate(name=payload.name)
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not site:
        raise HTTPException(status_code=404, detail="Site introuvable")
    cells = ["" for _ in range(364)]
    for cell in site.cells:
        if 0 <= cell.day_index < 364 and cell.value:
            cells[cell.day_index] = cell.value
    return WorkloadPlanSiteResponse(
        id=site.id,
        name=site.name,
        position=site.position,
        cells=cells,
    )


@app.delete("/api/workload-plan/sites/{site_id}", status_code=204)
def delete_workload_plan_site(
    _current_user: CurrentUser,
    site_id: int,
    session: Session = Depends(get_session),
) -> Response:
    deleted = crud.delete_workload_site(session, site_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Site introuvable")
    return Response(status_code=204)


@app.post("/api/workload-plan/cells")
def update_workload_plan_cells(
    _current_user: CurrentUser,
    payload: WorkloadPlanCellsPayload,
    session: Session = Depends(get_session),
) -> Dict[str, int]:
    updates = [
        WorkloadCellUpdate(**item.model_dump()) for item in payload.updates or []
    ]
    try:
        count = crud.bulk_update_workload_cells(session, updates)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"updated": count}


@app.post("/api/workload-plan/import")
def import_workload_plan(
    _current_user: CurrentUser,
    payload: WorkloadPlanImportPayload,
    session: Session = Depends(get_session),
) -> Dict[str, int]:
    try:
        count = crud.replace_workload_plan(session, payload.sites, payload.cells)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"sites": count}


@app.get("/api/workload-plan/export/excel")
def export_workload_plan_excel(
    _current_user: CurrentUser,
    session: Session = Depends(get_session),
) -> Response:
    sites = crud.list_workload_sites(session)
    buffer = _build_workload_plan_workbook(sites)
    return _template_response(buffer, "plan_de_charge.xlsx")


@app.post("/api/workload-plan/import/excel")
async def import_workload_plan_excel(
    _current_user: CurrentUser,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> Dict[str, int]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Aucun fichier sélectionné.")
    if not file.filename.lower().endswith(ALLOWED_IMPORT_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail="Format de fichier non supporté. Merci d'utiliser un fichier Excel (.xlsx).",
        )
    content = await file.read()
    try:
        sites, cells = parse_workload_plan_excel(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    count = crud.replace_workload_plan(session, sites, cells)
    return {"sites": count}


@app.post("/filtres-courroies/filtres")
async def create_filter_line(
    _current_user: CurrentUser,
    site: str = Form(...),
    equipment: str = Form(...),
    efficiency: Optional[str] = Form(None),
    format_type: str = Form(...),
    dimensions: Optional[str] = Form(None),
    quantity: int = Form(...),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    if format_type not in FILTER_FORMAT_LABELS:
        raise HTTPException(status_code=400, detail="Format de filtre invalide")

    if quantity < 1:
        raise HTTPException(status_code=400, detail="Quantité invalide")

    payload = FilterLineCreate(
        site=site.strip(),
        equipment=equipment.strip(),
        efficiency=(efficiency.strip() if efficiency else None),
        format_type=format_type,
        dimensions=(dimensions.strip() if dimensions else None),
        quantity=quantity,
        order_week=(order_week.strip().upper() if order_week else None),
    )
    crud.create_filter_line(session, payload)
    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/filtres-courroies/filtres/{line_id}/update")
async def update_filter_line(
    _current_user: CurrentUser,
    line_id: int,
    site: str = Form(...),
    equipment: str = Form(...),
    efficiency: Optional[str] = Form(None),
    format_type: str = Form(...),
    dimensions: Optional[str] = Form(None),
    quantity: int = Form(...),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    if format_type not in FILTER_FORMAT_LABELS:
        raise HTTPException(status_code=400, detail="Format de filtre invalide")

    if quantity < 1:
        raise HTTPException(status_code=400, detail="Quantité invalide")

    payload = FilterLineUpdate(
        site=site.strip(),
        equipment=equipment.strip(),
        efficiency=(efficiency.strip() if efficiency else None),
        format_type=format_type,
        dimensions=(dimensions.strip() if dimensions else None),
        quantity=quantity,
        order_week=(order_week.strip() if order_week else None),
    )

    updated = crud.update_filter_line(session, line_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Ligne filtre introuvable")

    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/filtres-courroies/filtres/import")
async def import_filter_lines(
    _current_user: CurrentUser,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    if not file.filename:
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=["Aucun fichier sélectionné."],
            filename="Import filtres",
            singular_label="ligne filtre",
            plural_label="lignes filtre",
        )

    if not file.filename.lower().endswith(ALLOWED_IMPORT_EXTENSIONS):
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=[
                "Format de fichier non supporté. Merci d'utiliser un fichier Excel (.xlsx)."
            ],
            filename=file.filename,
            singular_label="ligne filtre",
            plural_label="lignes filtre",
        )

    content = await file.read()
    try:
        rows = parse_filter_lines_excel(content)
    except ValueError as exc:
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=[str(exc)],
            filename=file.filename,
            singular_label="ligne filtre",
            plural_label="lignes filtre",
        )

    created = 0
    errors: List[str] = []
    for idx, row in enumerate(rows, start=1):
        payload = {
            key: value
            for key, value in row.items()
            if not key.startswith("__")
        }
        row_number = row.get("__row__", idx)
        try:
            crud.create_filter_line(session, FilterLineCreate(**payload))
            created += 1
        except Exception as exc:
            session.rollback()
            errors.append(f"Ligne {row_number} : {exc}")

    return _store_import_report(
        "/filtres-courroies",
        created=created,
        total=len(rows),
        errors=errors,
        filename=file.filename,
        singular_label="ligne filtre",
        plural_label="lignes filtre",
    )


@app.post("/filtres-courroies/filtres/{line_id}/delete")
def delete_filter_line(
    _current_user: CurrentUser,
    line_id: int,
    session: Session = Depends(get_session),
):
    if not crud.delete_filter_line(session, line_id):
        raise HTTPException(status_code=404, detail="Ligne filtre introuvable")
    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/filtres-courroies/courroies")
async def create_belt_line(
    _current_user: CurrentUser,
    site: str = Form(...),
    equipment: str = Form(...),
    reference: str = Form(...),
    quantity: int = Form(...),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    payload = BeltLineCreate(
        site=site.strip(),
        equipment=equipment.strip(),
        reference=reference.strip(),
        quantity=quantity,
        order_week=(order_week.strip().upper() if order_week else None),
    )
    crud.create_belt_line(session, payload)
    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/filtres-courroies/courroies/{line_id}/update")
async def update_belt_line(
    _current_user: CurrentUser,
    line_id: int,
    site: str = Form(...),
    equipment: str = Form(...),
    reference: str = Form(...),
    quantity: int = Form(...),
    order_week: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    if quantity < 1:
        raise HTTPException(status_code=400, detail="Quantité invalide")

    payload = BeltLineUpdate(
        site=site.strip(),
        equipment=equipment.strip(),
        reference=reference.strip(),
        quantity=quantity,
        order_week=(order_week.strip() if order_week else None),
    )

    updated = crud.update_belt_line(session, line_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Ligne courroie introuvable")

    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/filtres-courroies/courroies/import")
async def import_belt_lines(
    _current_user: CurrentUser,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    if not file.filename:
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=["Aucun fichier sélectionné."],
            filename="Import courroies",
            singular_label="ligne courroie",
            plural_label="lignes courroie",
        )

    if not file.filename.lower().endswith(ALLOWED_IMPORT_EXTENSIONS):
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=[
                "Format de fichier non supporté. Merci d'utiliser un fichier Excel (.xlsx)."
            ],
            filename=file.filename,
            singular_label="ligne courroie",
            plural_label="lignes courroie",
        )

    content = await file.read()
    try:
        rows = parse_belt_lines_excel(content)
    except ValueError as exc:
        return _store_import_report(
            "/filtres-courroies",
            created=0,
            total=0,
            errors=[str(exc)],
            filename=file.filename,
            singular_label="ligne courroie",
            plural_label="lignes courroie",
        )

    created = 0
    errors: List[str] = []
    for idx, row in enumerate(rows, start=1):
        payload = {
            key: value
            for key, value in row.items()
            if not key.startswith("__")
        }
        row_number = row.get("__row__", idx)
        try:
            crud.create_belt_line(session, BeltLineCreate(**payload))
            created += 1
        except Exception as exc:
            session.rollback()
            errors.append(f"Ligne {row_number} : {exc}")

    return _store_import_report(
        "/filtres-courroies",
        created=created,
        total=len(rows),
        errors=errors,
        filename=file.filename,
        singular_label="ligne courroie",
        plural_label="lignes courroie",
    )


@app.post("/filtres-courroies/courroies/{line_id}/delete")
def delete_belt_line(
    _current_user: CurrentUser,
    line_id: int,
    session: Session = Depends(get_session),
):
    if not crud.delete_belt_line(session, line_id):
        raise HTTPException(status_code=404, detail="Ligne courroie introuvable")
    return RedirectResponse("/filtres-courroies", status_code=303)


@app.post("/clients/import")
async def import_clients(
    _current_user: CurrentUser,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    if not file.filename:
        return _store_import_report(
            "/",
            created=0,
            total=0,
            errors=["Aucun fichier sélectionné."],
            filename="Import",
            singular_label="client",
            plural_label="clients",
        )

    if not file.filename.lower().endswith(ALLOWED_IMPORT_EXTENSIONS):
        return _store_import_report(
            "/",
            created=0,
            total=0,
            errors=[
                "Format de fichier non supporté. Merci d'utiliser un fichier Excel (.xlsx)."
            ],
            filename=file.filename,
            singular_label="client",
            plural_label="clients",
        )

    content = await file.read()
    try:
        rows = parse_clients_excel(content)
    except ValueError as exc:
        return _store_import_report(
            "/",
            created=0,
            total=0,
            errors=[str(exc)],
            filename=file.filename,
            singular_label="client",
            plural_label="clients",
        )

    created = 0
    errors: list[str] = []
    for idx, row in enumerate(rows, start=1):
        contacts_raw = row.get("contacts", [])
        contacts_payload = [ContactCreate(**contact) for contact in contacts_raw]
        payload = {
            k: v
            for k, v in row.items()
            if not k.startswith("__") and k != "contacts"
        }
        row_number = row.get("__row__", idx)
        try:
            company_name = payload.pop("company_name", None)
            if not company_name:
                raise ValueError("Nom d'entreprise manquant")
            status_value = payload.get("status")
            try:
                statut_bool = _status_to_bool(status_value) if status_value else None
            except HTTPException as exc:
                raise ValueError(str(exc.detail))
            entreprise = crud.ensure_entreprise(
                session,
                name=company_name,
                adresse_facturation=payload.get("billing_address"),
                tag=payload.get("tags"),
                statut=statut_bool,
            )
            payload["entreprise_id"] = entreprise.id
            crud.create_client(
                session,
                ClientCreate(**payload),
                contacts=contacts_payload,
            )
            created += 1
        except Exception as exc:
            session.rollback()
            errors.append(f"Ligne {row_number} : {exc}")

    return _store_import_report(
        "/",
        created=created,
        total=len(rows),
        errors=errors,
        filename=file.filename,
        singular_label="client",
        plural_label="clients",
    )


def _autofit_sheet(sheet, padding: int = 2, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = min(max_length + padding, max_width)


def _build_workload_plan_workbook(sites: Iterable) -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Plan de charge"

    headers = ["Site"] + [f"Jour {index + 1}" for index in range(364)]
    sheet.append(headers)

    def _export_value(value: Optional[str]) -> Optional[Union[int, str]]:
        if value == "bad":
            return 8
        if value == "warn":
            return 4
        return value

    for site in sites:
        values: List[Optional[Union[int, str]]] = [""] * 364
        for cell in getattr(site, "cells", []) or []:
            if cell and 0 <= getattr(cell, "day_index", -1) < 364 and cell.value:
                values[cell.day_index] = _export_value(cell.value)
        sheet.append([getattr(site, "name", "")] + values)

    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 32

    legend = workbook.create_sheet("Légende")
    legend.append(["Valeur", "Signification"])
    legend.append(["4", "Orange — intervention à confirmer (4 h)"])
    legend.append(["8", "Rouge — charge pleine (8 h)"])
    legend.append(["ok:4", "Vert 4 h — retour au vert depuis orange"])
    legend.append(["ok:8", "Vert 8 h — retour au vert depuis rouge"])
    legend.append(["ok", "Vert — intervention validée"])
    legend.append(["(vide)", "Aucune information planifiée pour ce jour."])
    legend.freeze_panes = "A2"
    _autofit_sheet(legend, padding=4, max_width=55)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_client_import_template() -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Clients"

    headers = [
        "company_name",
        "name",
        "email",
        "phone",
        "technician_name",
        "billing_address",
        "depannage",
        "astreinte",
        "tags",
        "status",
        "contact_2_name",
        "contact_2_email",
        "contact_2_phone",
        "contact_3_name",
        "contact_3_email",
        "contact_3_phone",
    ]
    sheet.append(headers)

    sample_rows = [
        {
            "company_name": "Exemple SARL",
            "name": "Alice Martin",
            "email": "alice@example.fr",
            "phone": "0102030405",
            "technician_name": "Léo Fournier",
            "billing_address": "10 rue des Fleurs\n75000 Paris",
            "depannage": "refacturable",
            "astreinte": "incluse_refacturable",
            "tags": "premium, 2024",
            "status": "actif",
            "contact_2_name": "Paul Martin",
            "contact_2_email": "paul@example.fr",
            "contact_2_phone": "0188776655",
        },
        {
            "company_name": "Solutions BTP",
            "name": "Bruno Carrel",
            "email": "bruno@solutionsbtp.fr",
            "phone": "0611223344",
            "technician_name": "Nina Perrot",
            "billing_address": "5 avenue du Port\n44000 Nantes",
            "depannage": "non_refacturable",
            "astreinte": "incluse_non_refacturable",
            "tags": "chantier",
            "status": "actif",
            "contact_2_name": "Sophie Lemaitre",
            "contact_2_email": "sophie@solutionsbtp.fr",
            "contact_3_name": "Service Achat",
            "contact_3_phone": "0245789650",
        },
        {
            "company_name": "Collectif Horizon",
            "name": "Chloé Bernard",
            "email": "chloe@collectif-horizon.fr",
            "phone": "0499887766",
            "technician_name": "",
            "billing_address": "42 boulevard National\n13001 Marseille",
            "depannage": "refacturable",
            "astreinte": "pas_d_astreinte",
            "tags": "association",
            "status": "inactif",
        },
    ]

    for row in sample_rows:
        sheet.append([row.get(column, "") for column in headers])

    sheet.freeze_panes = "A2"
    _autofit_sheet(sheet, padding=2, max_width=50)

    options_sheet = workbook.create_sheet("Options")
    options_sheet.append(["Champ", "Valeurs autorisées", "Description"])
    options_sheet.freeze_panes = "A2"

    def _format_options(options: Dict[str, str]) -> str:
        return "\n".join(f"{key} — {label}" for key, label in options.items())

    options_sheet.append([
        "depannage",
        _format_options(DEPANNAGE_OPTIONS),
        "Détermine si les interventions sont refacturées.",
    ])
    options_sheet.append([
        "astreinte",
        _format_options(ASTREINTE_OPTIONS),
        "Choisissez le type d'astreinte applicable.",
    ])
    options_sheet.append([
        "technician_name",
        "Nom complet",
        "Identifie le technicien référent pour ce client.",
    ])
    options_sheet.append([
        "status",
        _format_options(STATUS_OPTIONS),
        "Etat du client dans votre CRM.",
    ])
    options_sheet.append([
        "contacts supplémentaires",
        "contact_2_name, contact_2_email, contact_2_phone, contact_3_name…",
        "Dupliquez le numéro (contact_4_*, contact_5_* …) pour ajouter des contacts additionnels. Seul le nom est obligatoire.",
    ])

    _autofit_sheet(options_sheet, padding=4, max_width=70)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_prestation_import_template() -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Prestations"

    headers = [
        "company_name",
        "client_name",
        "client_id",
        "prestation",
        "prestation_label",
        "budget",
        "frequency",
        "frequency_interval",
        "frequency_unit",
        "status",
        "order_week",
        "realization_week",
    ]
    sheet.append(headers)

    sample_rows = [
        {
            "company_name": "Exemple SARL",
            "client_name": "Alice Martin",
            "prestation": "controle_ssi",
            "budget": "2500",
            "frequency": "contrat_annuel",
            "status": "en_cours",
            "order_week": "S05",
        },
        {
            "company_name": "Solutions BTP",
            "client_name": "Bruno Carrel",
            "prestation": "maintenance_groupe_electrogene",
            "budget": "4800",
            "frequency": "contrat_trimestriel",
            "status": "non_commence",
        },
        {
            "company_name": "Collectif Horizon",
            "client_name": "Chloé Bernard",
            "prestation_label": "Analyse d'eau",
            "frequency": "prestation_ponctuelle",
            "status": "fait",
            "realization_week": "S12",
        },
    ]

    for row in sample_rows:
        sheet.append([row.get(column, "") for column in headers])

    sheet.freeze_panes = "A2"
    _autofit_sheet(sheet, padding=2, max_width=40)

    options_sheet = workbook.create_sheet("Options")
    options_sheet.append(["Champ", "Valeurs autorisées", "Description"])
    options_sheet.freeze_panes = "A2"

    status_values = "\n".join(
        f"{key} — {label}" for key, label in SUBCONTRACT_STATUS_OPTIONS.items()
    )
    frequency_values = "\n".join(
        f"{key} — {data['label']}" for key, data in PREDEFINED_FREQUENCIES.items()
    )

    options_sheet.append(
        [
            "prestation",
            "Clé interne du référentiel",
            "Utilisez la colonne Prestation ou renseignez uniquement le libellé dans prestation_label.",
        ]
    )
    options_sheet.append(
        [
            "frequency",
            frequency_values,
            "Utilisez custom_interval ou un format interval:unite:valeur pour les fréquences personnalisées.",
        ]
    )
    options_sheet.append(
        [
            "status",
            status_values,
            "Statut opérationnel appliqué à la ligne importée.",
        ]
    )
    options_sheet.append(
        [
            "frequency_unit",
            "months — mois\nyears — années",
            "Uniquement requis pour les fréquences personnalisées.",
        ]
    )

    _autofit_sheet(options_sheet, padding=4, max_width=70)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_prestation_reference_export(
    definitions: Iterable[PrestationDefinition],
) -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Référentiel"

    headers = [
        "Libellé prestation",
        "Code budget",
        "Catégorie",
    ]
    sheet.append(headers)

    sorted_definitions = sorted(
        definitions,
        key=lambda definition: (
            getattr(definition, "category", "") or "",
            getattr(definition, "label", "") or "",
        ),
    )

    for definition in sorted_definitions:
        sheet.append(
            [
                getattr(definition, "label", ""),
                getattr(definition, "budget_code", ""),
                getattr(definition, "category", ""),
            ]
        )

    sheet.freeze_panes = "A2"
    _autofit_sheet(sheet, padding=2, max_width=45)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_filter_import_template() -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Filtres"

    headers = [
        "site",
        "equipment",
        "format_type",
        "efficiency",
        "dimensions",
        "quantity",
        "order_week",
    ]
    sheet.append(headers)

    sample_rows = [
        {
            "site": "Data Center Nord",
            "equipment": "CTA N°1",
            "format_type": "cadre",
            "efficiency": "ISO ePM1 80%",
            "dimensions": "592 x 592 x 47",
            "quantity": 3,
            "order_week": "S12",
        },
        {
            "site": "Siège social",
            "equipment": "Unité Rooftop",
            "format_type": "poche",
            "efficiency": "F7",
            "dimensions": "287 x 592 x 635",
            "quantity": 2,
            "order_week": "S22",
        },
    ]

    for row in sample_rows:
        sheet.append([row.get(column, "") for column in headers])

    sheet.freeze_panes = "A2"
    _autofit_sheet(sheet, padding=2, max_width=55)

    options_sheet = workbook.create_sheet("Options")
    options_sheet.append(["format_type", "Libellé"])
    options_sheet.freeze_panes = "A2"

    for value, label in FILTER_FORMAT_OPTIONS:
        options_sheet.append([value, label])

    _autofit_sheet(options_sheet, padding=4, max_width=50)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_belt_import_template() -> BytesIO:
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Courroies"

    headers = ["site", "equipment", "reference", "quantity", "order_week"]
    sheet.append(headers)

    sample_rows = [
        {
            "site": "Usine Est",
            "equipment": "Compresseur A",
            "reference": "SPB-2000",
            "quantity": 2,
            "order_week": "S08",
        },
        {
            "site": "Usine Est",
            "equipment": "Ventilateur extraction",
            "reference": "XPZ-1250",
            "quantity": 3,
            "order_week": "S30",
        },
        {
            "site": "Entrepôt Sud",
            "equipment": "CTA zone picking",
            "reference": "SPC-1780",
            "quantity": 1,
            "order_week": "",
        },
    ]

    for row in sample_rows:
        sheet.append([row.get(column, "") for column in headers])

    sheet.freeze_panes = "A2"
    _autofit_sheet(sheet, padding=2, max_width=55)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _template_response(buffer: BytesIO, filename: str) -> Response:
    content = buffer.getvalue()
    headers_dict = {
        "Content-Disposition": (
            f"attachment; filename={filename}; "
            f"filename*=UTF-8''{filename}"
        )
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_dict,
    )


@app.get("/clients/import/template")
@app.get("/clients/import/template/")
def download_client_import_template(_current_user: CurrentUser):
    return _template_response(
        _build_client_import_template(), "modele_import_clients.xlsx"
    )


@app.get("/prestations/import/template")
@app.get("/prestations/import/template/")
def download_prestation_import_template(_current_user: CurrentUser):
    return _template_response(
        _build_prestation_import_template(), "modele_import_prestations.xlsx"
    )


@app.get(
    "/prestations/referentiel/export",
    name="download_prestation_reference",
)
@app.get("/prestations/referentiel/export/", include_in_schema=False)
def download_prestation_reference(
    _current_user: CurrentUser, session: Session = Depends(get_session)
):
    definitions = crud.list_prestation_definitions(session)
    return _template_response(
        _build_prestation_reference_export(definitions),
        "referentiel_prestations.xlsx",
    )


@app.get("/filtres-courroies/filtres/import/template")
@app.get("/filtres-courroies/filtres/import/template/")
def download_filter_import_template(_current_user: CurrentUser):
    return _template_response(
        _build_filter_import_template(), "modele_import_filtres.xlsx"
    )


@app.get("/filtres-courroies/courroies/import/template")
@app.get("/filtres-courroies/courroies/import/template/")
def download_belt_import_template(_current_user: CurrentUser):
    return _template_response(
        _build_belt_import_template(), "modele_import_courroies.xlsx"
    )
